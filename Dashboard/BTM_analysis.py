import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook


# st.set_page_config(page_title="BTM Analysis", layout="wide")
st.set_page_config(
    page_title="Segment & Brand Performance Dashboard",
    page_icon=r"..\Image\perno_logo.png",
    layout="wide"
)

from PIL import Image
import streamlit as st



# Load logo from your local path
logo_path = r"..\Image\perno_logo.png"   # ✅ raw string avoids path errors
logo = Image.open(logo_path)

# Header layout: Logo + Title
col1, col2 = st.columns([1, 14])

with col1:
    st.image(logo, width=100)

with col2:
    st.markdown(
        """
        <h1 style="margin-bottom:0;">Segment & Brand Performance Dashboard</h1>
        <p style="color:gray; margin-top:-10px;">Pernod Ricard India</p>
        """,
        unsafe_allow_html=True
    )

# st.title("📂 BTM Analysis")
# Apply full-width tabs
st.markdown(
        """
        <style>
            div.stTabs button {
                flex-grow: 1;
                text-align: center;
            }
        </style>
        """,
        unsafe_allow_html=True
    )

st.markdown("""
<style>
/* --- ONLY increase TAB TEXT, NOT button size --- */
div[data-testid="stTabs"] button p {
    font-size: 20px !important;
    font-weight: 800 !important;
    margin: 0 !important;
}

/* --- Active tab emphasis only on TEXT --- */
div[data-testid="stTabs"] button[aria-selected="true"] p {
    font-size: 24px !important;
    font-weight: 800 !important;
    color: red !important;
}
</style>
""", unsafe_allow_html=True)




tab1, tab2, tab3 = st.tabs(["BTM Summary", "Maestria", "Dashboard"])

with tab1:

    # --- Cache Data Loading ---
    @st.cache_data
    def load_data(file):
        df = pd.read_excel(file, sheet_name="Data")
        df = df.dropna(axis=1, how='all')
        return df

    # --- Define Fiscal Month to Quarter Mapping ---
    fiscal_quarter_map = {
        'July': 'Q1', 'August': 'Q1', 'September': 'Q1',
        'October': 'Q2', 'November': 'Q2', 'December': 'Q2',
        'January': 'Q3', 'February': 'Q3', 'March': 'Q3',
        'April': 'Q4', 'May': 'Q4', 'June': 'Q4'
    }
    fiscal_months = list(fiscal_quarter_map.keys())

    # --- File Upload ---
    uploaded_file1 = st.sidebar.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

    if uploaded_file1 is not None:
        try:
            df = load_data(uploaded_file1)
            st.success("✅ File loaded successfully (cached for faster reloads)")

            # --- Check Required Columns ---
            required_cols = ['Brand', 'PRI Year', 'Month', 'NS M INR', 'Revised Seg']
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                st.error(f"❌ Missing columns: {missing}")
            else:
                # --- Clean Month Order & Add Quarter ---
                df['Month'] = pd.Categorical(df['Month'], categories=fiscal_months, ordered=True)
                df['Quarter'] = df['Month'].map(fiscal_quarter_map)

                # --- Filter by Revised Segments (multiple) ---
                seg_list = sorted(df['Revised Seg'].dropna().unique())
                selected_seg = st.multiselect(
                    "Select Revised Segment(s):",
                    options=seg_list,
                    default=[],
                    #default=seg_list[:1],  # optional, default selection
                )

                if not selected_seg:
                    st.warning("Please select at least one segment to continue.")
                    st.stop()

                count_brand_family = df.groupby("Revised Seg")["Brand Family"].nunique().reset_index()
                count_brand_family.columns = ["Revised Seg", "Unique Brand Family Count"]
                # st.dataframe(count_brand_family)


                #### data filtering based on segment selection ####    

                df_filtered = df[df['Revised Seg'].isin(selected_seg)]
                df_filtered = df_filtered[df_filtered['PRI Year']!="A26"]
                df_filtered['State'] = df_filtered['State'].str.title()

                # --- Brand Family Multiselect ---
                brand_family_opts = sorted(df_filtered["Brand Family"].dropna().unique().tolist())  # options
                # st.write(brand_family_opts)

           



                df_filtered_zone = df_filtered.copy()
                # st.dataframe(df_filtered)

                # --- Manufacturing Pivot: Absolute Numbers ---
                pivot_mfs = pd.pivot_table(
                    df_filtered,
                    index='Mfg Com',
                    columns='PRI Year',
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                ).sort_index(axis=1)
                

                # --- Filter only A24 and A25 ---
                pivot_mfs = pivot_mfs[[col for col in pivot_mfs.columns if col in ['A23','A24', 'A25']]]
                # pivot_mfs = pivot_mfs.round(0).astype(int)
                pivot_mfs_fmt = pivot_mfs.copy().applymap(lambda x: f"{x:.0f}" if pd.notnull(x) else "")

                # --- Year-over-Year Growth (%) ---
                pivot_mfs_growth = pivot_mfs.pct_change(axis=1) * 100
                pivot_mfs_growth = pivot_mfs_growth.round(2)
                pivot_mfs_growth = pivot_mfs_growth[[col for col in pivot_mfs_growth.columns if col in ['A24', 'A25']]]
                pivot_mfs_growth.rename(columns=lambda x: f"{x} Growth %", inplace=True)
                pivot_mfs_growth_fmt = pivot_mfs_growth.copy().applymap(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")

                # # --- Streamlit Display ---
                # st.subheader(f"🏭 Manufacturing Company — Absolute NS (₹ Mn) — {selected_seg}")
                # st.dataframe(pivot_mfs_fmt)

                # st.subheader(f"📈 Manufacturing Company — YoY NS Growth % (A24 → A25) — {selected_seg}")
                # st.dataframe(pivot_mfs_growth_fmt)

                # ---- display Brand × Mfg Com × Year with Growth % ----
                st.subheader(f"🏭 Manufacturing Co. View — {selected_seg}")
                final_mfs_review = pd.concat(
                    [pivot_mfs_fmt, pivot_mfs_growth_fmt],
                    axis=1
                )
                final_mfs_review = final_mfs_review.reset_index()
                # st.write(final_mfs_review.columns)
                # --- Custom order for Mfg Com ---
                custom_order = ['PRI', 'Diageo', 'Others']
                final_mfs_review['Mfg Com'] = pd.Categorical(final_mfs_review['Mfg Com'], categories=custom_order, ordered=True)
                final_mfs_review = final_mfs_review.sort_values(by='Mfg Com').reset_index(drop=True)
                st.dataframe(final_mfs_review)

                # -------------------------------------------------------------
                # 📝 ADD COMMENT BOX BELOW THE TABLE (Shared Across All Tabs)
                # -------------------------------------------------------------

                # Create storage if not exists
                if "mfg_table_comment" not in st.session_state:
                    st.session_state["mfg_table_comment"] = ""

                # Text box that stores comment globally
                comment = st.text_area(
                    "📝 Comment on Manufacturing Company View:",
                    value=st.session_state["mfg_table_comment"],
                    key="mfg_review_comment_box"
                )

                # Update session state when user edits the comment
                st.session_state["mfg_table_comment"] = comment



                # --- Pivot: Brand × Manufacturing Company × Year ---
                pivot_brand_mfs = pd.pivot_table(
                    df_filtered,
                    index=['Brand', 'Mfg Com'],
                    columns='PRI Year',
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                ).sort_index(axis=1)

                # --- Keep only A23, A24, A25 ---
                years = ['A23', 'A24', 'A25']
                pivot_brand_mfs = pivot_brand_mfs[[y for y in years if y in pivot_brand_mfs.columns]]

                # --- Compute YoY Growth (%) for A24 and A25 ---
                pivot_brand_mfs_growth = pivot_brand_mfs.pct_change(axis=1) * 100
                pivot_brand_mfs_growth = pivot_brand_mfs_growth.round(2)

                # --- Combine absolute + growth ---
                combined_df = pivot_brand_mfs.copy()
                for year in ['A24', 'A25']:
                    if year in pivot_brand_mfs_growth.columns:
                        combined_df[f"{year} Growth %"] = pivot_brand_mfs_growth[year]

                # --- Flatten Index ---
                combined_df = combined_df.reset_index()

                # --- Reorder Columns ---
                final_cols = (
                    ['Mfg Com', 'Brand']
                    + [y for y in ['A23', 'A24', 'A25'] if y in combined_df.columns]
                    + [f"{y} Growth %" for y in ['A24', 'A25'] if f"{y} Growth %" in combined_df.columns]
                )
                combined_df = combined_df[final_cols]

                # st.dataframe(combined_df)

                # --- Custom order for Mfg Com ---
                custom_order = ['PRI', 'Diageo', 'Others']
                combined_df['Mfg Com'] = pd.Categorical(combined_df['Mfg Com'], categories=custom_order, ordered=True)
                combined_df = combined_df.sort_values(by='Mfg Com').reset_index(drop=True)

                # --- Add one Segment Total (sum across all rows) ---
                segment_total = (
                    combined_df[[c for c in ['A23', 'A24', 'A25'] if c in combined_df.columns]]
                    .sum()
                    .to_frame()
                    .T
                )
                segment_total['Mfg Com'] = 'All Mfg Com'
                segment_total['Brand'] = 'Segment Total'

                # --- Compute Brand Totals (sum across all Mfg Com for each Brand) ---
                brand_totals = (
                    combined_df.groupby('Mfg Com', dropna=False)
                    [[c for c in ['A23', 'A24', 'A25'] if c in combined_df.columns]]
                    .sum()
                    .reset_index()
                )
                brand_totals['Brand'] = 'Brand Total'

                # --- Combine all: add brand totals and segment total ---
                combined_with_total = pd.concat([combined_df, brand_totals, segment_total], ignore_index=True)

                combined_with_total["A24 Growth %"] = (combined_with_total["A24"] - combined_with_total["A23"])/combined_with_total["A23"] * 100
                combined_with_total["A25 Growth %"] = (combined_with_total["A25"] - combined_with_total["A24"])/combined_with_total["A24"] * 100

                # --- Sort: Brand Total appears first within each Brand; Segment Total appears first overall ---
                combined_with_total['SortOrder'] = 0
                combined_with_total.loc[combined_with_total['Brand'] == 'Segment Total', 'SortOrder'] = 2
                combined_with_total.loc[combined_with_total['Mfg Com'] == 'Brand Total', 'SortOrder'] = 1

                combined_with_total = combined_with_total.sort_values(
                    by=['SortOrder', 'Brand', 'A25'],
                    ascending=[False, True, False],
                    ignore_index=True
                )

                combined_with_total = combined_with_total.drop(columns='SortOrder')

                # --- Add A25 Brand Share (%) ---
                if 'A25' in combined_with_total.columns:
                    total_a25 = float(segment_total['A25'])
                    combined_with_total['A25 Brand Share %'] = (
                        combined_with_total['A25'] / total_a25 * 100
                    ).round(2)
                    combined_with_total.loc[
                        combined_with_total['Brand'] == 'Segment Total', 'A25 Brand Share %'
                    ] = 100

                # --- Custom order for Mfg Com ---
                custom_order = ['PRI', 'Diageo', 'Others']

                # Apply categorical sorting for Mfg Com
                combined_with_total['Mfg Com'] = pd.Categorical(
                    combined_with_total['Mfg Com'], 
                    categories=custom_order, 
                    ordered=True
                )

                # --- Sort brands within Mfg Com by A25 descending (except Segment Total) ---
                if 'A25' in combined_with_total.columns:
                    combined_with_total = combined_with_total.sort_values(
                        by=['Mfg Com', 'A25'],
                        ascending=[True, False],
                        ignore_index=True
                    )

                # --- Ensure Segment Total row is always first ---
                segment_total_row = combined_with_total[combined_with_total['Brand'] == 'Segment Total']
                other_rows = combined_with_total[combined_with_total['Brand'] != 'Segment Total']

                # Combine — Segment Total first, then custom order
                all_seg_brand_final_df = pd.concat([segment_total_row, other_rows], ignore_index=True)

                # st.dataframe(all_seg_brand_final_df)
                # --- Copy working df ---
                display_df = all_seg_brand_final_df.copy()

                # --- Custom order for Mfg Com ---
                custom_order = ['PRI', 'Diageo', 'Others']
                display_df['Mfg Com'] = pd.Categorical(display_df['Mfg Com'], categories=custom_order, ordered=True)

                # --- Sort brands within each Mfg Com by A25 descending ---
                if 'A25' in display_df.columns:
                    display_df = display_df.sort_values(by=['Mfg Com', 'A25'], ascending=[True, False], ignore_index=True)
                else:
                    st.warning("⚠️ 'A25' column not found; skipping sort.")

                # --- Store brand order for consistency ---
                brand_order = display_df['Brand'].tolist()

                # --- Move Segment Total row to top ---
                segment_total_row = display_df[display_df['Brand'] == 'Segment Total']
                other_rows = display_df[display_df['Brand'] != 'Segment Total']
                display_df = pd.concat([segment_total_row, other_rows], ignore_index=True)

                # --- Prepare formatted copy for export ---
                export_display_df = display_df.copy()

                # Format A23, A24, A25 numeric columns (1 decimal)
                for col in ['A23', 'A24', 'A25']:
                    if col in export_display_df.columns:
                        export_display_df[col] = export_display_df[col].apply(
                            lambda x: f"{x:,.1f}" if pd.notna(x) else "-"
                        )

                # Format Growth columns (2 decimals + % sign)
                for col in export_display_df.columns:
                    if "Growth" in col:
                        export_display_df[col] = export_display_df[col].apply(
                            lambda x: f"{x:.2f}%" if pd.notna(x) else "-"
                        )

                # Round remaining numeric columns to 2 decimals
                num_cols = export_display_df.select_dtypes(include="number").columns
                for col in num_cols:
                    if col not in ['A23', 'A24', 'A25'] and "Growth" not in col:
                        export_display_df[col] = export_display_df[col].round(2)

                # --- Streamlit Display (styled) ---
                styled_display = (
                    display_df.style
                        .format(na_rep="-", precision=2)
                        .format(subset=[c for c in display_df.columns if c in ['A23','A24','A25']], formatter="{:.0f}")
                        .format(subset=[c for c in display_df.columns if 'Growth' in c], formatter="{:.2f}%")
                        .format(subset=['A25 Brand Share %'], formatter="{:.2f}%")
                        # Optional: highlight Segment Total
                        .apply(lambda s: ['font-weight: bold; background-color: #f2f2f2' if s['Brand'] == 'Segment Total' else '' for _ in s], axis=1)
                        .apply(lambda s: ['font-weight: bold; background-color: #f9f1f1' if s['Brand'] == 'Brand Total' else '' for _ in s], axis=1)
                )

                # --- Streamlit Display ---
                st.subheader(f"🏭 Brand × Manufacturing Company — NS (₹ Mn) & YoY Growth — {selected_seg}")
                st.dataframe(styled_display)




                # initialize session variable if not exists
                if "selected_brand_family" not in st.session_state:
                    st.session_state["selected_brand_family"] = brand_family_opts  # default

        
                st.write("### Select Brand Families")
                selected_brand_family = st.multiselect(
                    "Include Brand Families",
                    options=brand_family_opts,
                    default=[],
                    #default=brand_family_opts,
                    key="brand_family_multiselect",
                    placeholder="Choose brand families…",
                )

                # keep updating session state dynamically with user changes
                st.session_state["selected_brand_family"] = selected_brand_family

                df_filtered = df_filtered[df_filtered['Brand Family'].isin(selected_brand_family)]
                # st.write(f"Filtered to {df_filtered['Brand'].unique()} Brands across {df_filtered['Brand Family'].unique()} Brand Families.")



                st.write("### Select Brands")

                # Group brands by brand family (filtered df_filtered assumed)

                # Step 2: Build brand options grouped by family, excluding families appearing as brands
                family_to_brands = {}
                for fam in selected_brand_family:
                    brands_in_family = sorted(
                        #df_filtered[(df_filtered["Brand Family"] == fam) & (df_filtered["Brand"] != fam)]["Brand"].dropna().unique().tolist()
                        df_filtered[(df_filtered["Brand Family"] == fam)]["Brand"].dropna().unique().tolist()
                    )
                    family_to_brands[fam] = brands_in_family

                # Step 3: Initialize session state's selected_brands *only once* when brand families change
                if "prev_selected_brand_family" not in st.session_state or st.session_state.prev_selected_brand_family != selected_brand_family:
                    all_brands = [brand for brands in family_to_brands.values() for brand in brands]
                    st.session_state.selected_brands = all_brands.copy()  # set all selected
                    st.session_state.prev_selected_brand_family = selected_brand_family  # remember current brand family

                # Step 4: Render checkboxes reflecting current selections
                selected_brands = []
                for family, brands in family_to_brands.items():
                    with st.expander(family, expanded=True):
                        for brand in brands:
                            checked = st.checkbox(
                                brand,
                                value=brand in st.session_state.selected_brands,
                                key=f"checkbox_{brand}"
                            )
                            if checked:
                                selected_brands.append(brand)

                # Update session state based on current selections
                st.session_state["selected_brands"] = selected_brands


                # After brand checkbox selection:
                if not st.session_state.get("selected_brands"):
                    st.warning("⚠️ Please select at least one Brand.")
                    st.stop()  # stop further code execution until user selects


                brand_order = [b for b in brand_order if b in selected_brands]
                # st.write(brand_order)

                

                df_filtered = df_filtered[df_filtered['Brand'].isin(selected_brands)]

                # --- Monthly Pivot ---
                pivot_monthly = pd.pivot_table(
                    df_filtered,
                    index=['Brand Family','Brand'],
                    columns=['PRI Year', 'Month'],
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                ).sort_index(axis=1, level=[0, 1])
                

                # --- Filter only A23–A25 years ---
                years_to_keep = ['A23', 'A24', 'A25']
                pivot_monthly_clean = pivot_monthly[[y for y in pivot_monthly.columns.levels[0] if y in years_to_keep]]
                

                # --- Remove rows where all values are zero ---
                pivot_monthly_clean_nonzero = pivot_monthly_clean.loc[
                    (pivot_monthly_clean != 0).any(axis=1)
                ]

                # --- Flatten the MultiIndex columns into "Year-Month" strings ---
                pivot_monthly_flat = pivot_monthly_clean_nonzero.copy()
                pivot_monthly_flat.columns = [f"{year}-{month}" for year, month in pivot_monthly_flat.columns]

                # --- Compute Brand Family Totals ---
                brand_family_totals = pivot_monthly_flat.groupby(level=0).sum()
                brand_family_totals['Brand'] = brand_family_totals.index + ' Total'

                # Recreate MultiIndex for totals
                brand_family_totals.index = pd.MultiIndex.from_arrays(
                    [brand_family_totals.index, brand_family_totals['Brand']],
                    names=['Brand Family', 'Brand']
                )
                brand_family_totals = brand_family_totals.drop(columns='Brand')

                # --- Combine totals with original pivot ---
                combined_pivot = pd.concat([pivot_monthly_flat, brand_family_totals])

                # --- Sort so totals come first for each Brand Family ---
                combined_pivot['SortOrder'] = combined_pivot.index.get_level_values('Brand').str.contains('Total').astype(int)
                combined_pivot = combined_pivot.sort_values(
                    by=['Brand Family', 'SortOrder', 'Brand'],
                    ascending=[True, False, True]
                ).drop(columns='SortOrder')

                # --- ✅ Add "All Segment Total" row at the top (based only on Brand Family Totals) ---
                brand_family_total_rows = combined_pivot.loc[
                    combined_pivot.index.get_level_values('Brand').str.contains('Total')
                ]

                segment_total = brand_family_total_rows.sum(numeric_only=True)
                segment_total.name = ('All Segment', 'All Segment Total')

                combined_pivot = pd.concat([
                    pd.DataFrame([segment_total],
                                index=pd.MultiIndex.from_tuples([segment_total.name],
                                                                names=['Brand Family', 'Brand'])),
                    combined_pivot
                ])


                # --- Convert numeric columns to float ---
                combined_pivot = combined_pivot.apply(pd.to_numeric, errors='coerce')

                # --- Monthly Volume Pivot ---
                pivot_monthly_vol = pd.pivot_table(
                    df_filtered,
                    index=['Brand Family', 'Brand'],
                    columns=['PRI Year', 'Month'],
                    values='Vol K CS',       # 👈 change here from 'NS M INR' to 'Volume'
                    aggfunc='sum',
                    fill_value=0
                ).sort_index(axis=1, level=[0, 1])


                # --- Filter only A23–A25 years ---
                years_to_keep = ['A23', 'A24', 'A25']
                pivot_monthly_vol_clean = pivot_monthly_vol[
                    [y for y in pivot_monthly_vol.columns.levels[0] if y in years_to_keep]
                ]


                # --- Remove rows where all values are zero ---
                pivot_monthly_vol_clean_nonzero = pivot_monthly_vol_clean.loc[
                    (pivot_monthly_vol_clean != 0).any(axis=1)
                ]


                # --- Flatten the MultiIndex columns into "Year-Month" strings ---
                pivot_monthly_vol_flat = pivot_monthly_vol_clean_nonzero.copy()
                pivot_monthly_vol_flat.columns = [
                    f"{year}-{month}" for year, month in pivot_monthly_vol_flat.columns
                ]


                # --- Compute Brand Family Totals ---
                brand_family_totals_vol = pivot_monthly_vol_flat.groupby(level=0).sum()
                brand_family_totals_vol['Brand'] = brand_family_totals_vol.index + ' Total'

                # Recreate MultiIndex for totals
                brand_family_totals_vol.index = pd.MultiIndex.from_arrays(
                    [brand_family_totals_vol.index, brand_family_totals_vol['Brand']],
                    names=['Brand Family', 'Brand']
                )
                brand_family_totals_vol = brand_family_totals_vol.drop(columns='Brand')


                # --- Combine totals with original pivot ---
                combined_pivot_vol = pd.concat([pivot_monthly_vol_flat, brand_family_totals_vol])

                # --- Sort so totals come first for each Brand Family ---
                combined_pivot_vol['SortOrder'] = combined_pivot_vol.index.get_level_values('Brand').str.contains('Total').astype(int)
                combined_pivot_vol = combined_pivot_vol.sort_values(
                    by=['Brand Family', 'SortOrder', 'Brand'],
                    ascending=[True, False, True]
                ).drop(columns='SortOrder')


                # --- ✅ Add "All Segment Total" row based only on Brand Family Totals ---
                brand_family_total_rows = combined_pivot_vol.loc[
                    combined_pivot_vol.index.get_level_values('Brand').str.contains('Total')
                ]

                segment_total_vol = brand_family_total_rows.sum(numeric_only=True)
                segment_total_vol.name = ('All Segment', 'All Segment Total')

                combined_pivot_vol = pd.concat([
                    pd.DataFrame([segment_total_vol],
                                index=pd.MultiIndex.from_tuples([segment_total_vol.name],
                                                                names=['Brand Family', 'Brand'])),
                    combined_pivot_vol
                ])

                # --- Convert numeric columns to float ---
                combined_pivot_vol = combined_pivot_vol.apply(pd.to_numeric, errors='coerce')






                # --- Step 1: Reset index for clarity ---
                combined_pivot_reset = combined_pivot.reset_index()

                # --- Step 2: Define fiscal month order ---
                fiscal_months = [
                    'July', 'August', 'September', 'October', 'November', 'December',
                    'January', 'February', 'March', 'April', 'May', 'June'
                ]

                # --- Step 3: Identify only year-month columns ---
                month_cols = [
                    c for c in combined_pivot_reset.columns
                    if isinstance(c, str) and '-' in c  # e.g. 'A23-July'
                ]

                # --- Step 4: Sort columns chronologically (fiscal order) ---
                sorted_month_cols = sorted(
                    month_cols,
                    key=lambda x: (x.split('-')[0], fiscal_months.index(x.split('-')[1]))
                )
                # st.dataframe(combined_pivot)

                # --- Step 5: Compute month-over-month growth (% change) ---
                monthly_growth_df = combined_pivot_reset.copy()

                for i, col in enumerate(sorted_month_cols):
                    if i == 0:
                        monthly_growth_df[col] = 0  # no growth for first month
                    else:
                        prev_col = sorted_month_cols[i - 1]
                        monthly_growth_df[col] = (
                            (combined_pivot_reset[col] - combined_pivot_reset[prev_col]) /
                            combined_pivot_reset[prev_col].replace(0, pd.NA)
                        ) * 100

                # --- Step 6: Order columns nicely ---
                ordered_cols = ['Brand Family', 'Brand'] + sorted_month_cols
                monthly_growth_df = monthly_growth_df[ordered_cols]

                # --- Step 7: Optional formatting ---
                monthly_growth_df = monthly_growth_df.round(2)

               

                # --- Quarterly Pivot ---
                pivot_quarterly = pd.pivot_table(
                    df_filtered,
                    index=['Brand Family', 'Brand'],
                    columns=['PRI Year', 'Quarter'],
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                ).sort_index(axis=1, level=[0, 1])

                # --- Filter only A23–A25 years ---
                years_to_keep = ['A23', 'A24', 'A25']
                pivot_quarterly_fil = pivot_quarterly[
                    [y for y in pivot_quarterly.columns.levels[0] if y in years_to_keep]
                ]

                # --- Map fiscal quarters to friendly labels ---
                quarter_custom_map = {
                    'Q1': 'JAS',   # July–Aug–Sep
                    'Q2': 'OND',   # Oct–Nov–Dec
                    'Q3': 'JFM',   # Jan–Feb–Mar
                    'Q4': 'AMJ'    # Apr–May–Jun
                }
                pivot_quarterly_fil.columns = pd.MultiIndex.from_tuples([
                    (year, quarter_custom_map.get(q, q))
                    for year, q in pivot_quarterly_fil.columns
                ])

                # --- Flatten MultiIndex columns to "A23_JAS" style ---
                pivot_quarterly_flat = pivot_quarterly_fil.copy()
                pivot_quarterly_flat.columns = [f"{year}_{q}" for year, q in pivot_quarterly_flat.columns]

                # --- Compute Brand Family Totals ---
                brand_family_totals = pivot_quarterly_flat.groupby(level=0).sum()
                brand_family_totals['Brand'] = brand_family_totals.index + ' Total'
                brand_family_totals.index = pd.MultiIndex.from_arrays(
                    [brand_family_totals.index, brand_family_totals['Brand']],
                    names=['Brand Family', 'Brand']
                )
                brand_family_totals = brand_family_totals.drop(columns='Brand')

                # --- Combine with main data ---
                combined_quarterly = pd.concat([pivot_quarterly_flat, brand_family_totals])

                # --- Sort: totals first within each Brand Family ---
                combined_quarterly['SortOrder'] = combined_quarterly.index.get_level_values('Brand').str.contains('Total').astype(int)
                combined_quarterly = combined_quarterly.sort_values(
                    by=['Brand Family', 'SortOrder', 'Brand'],
                    ascending=[True, False, True]
                ).drop(columns='SortOrder')

                # --- ✅ Add "All Segment Total" row (sum of all Brand Family Totals only) ---
                brand_family_total_rows = combined_quarterly.loc[
                    combined_quarterly.index.get_level_values('Brand').str.contains('Total')
                ]

                segment_total = brand_family_total_rows.sum(numeric_only=True)
                segment_total.name = ('All Segment', 'All Segment Total')

                # Make it a MultiIndex row and append at TOP
                combined_quarterly = pd.concat([
                    pd.DataFrame([segment_total],
                                index=pd.MultiIndex.from_tuples([segment_total.name],
                                                                names=['Brand Family', 'Brand'])),
                    combined_quarterly
                ])


                # --- Sort columns chronologically ---
                quarter_order = []
                for year in sorted({col.split("_")[0] for col in combined_quarterly.columns}):
                    for q in ['JAS', 'OND', 'JFM', 'AMJ']:
                        col_name = f"{year}_{q}"
                        if col_name in combined_quarterly.columns:
                            quarter_order.append(col_name)

                combined_quarterly = combined_quarterly[quarter_order]

                # --- Compute QoQ Growth (% change sequentially across quarters) ---
                quarterly_growth_qoq = combined_quarterly.copy()
                prev_values = None
                for col in quarter_order:
                    if prev_values is None:
                        quarterly_growth_qoq[col + " Growth %"] = pd.NA
                    else:
                        quarterly_growth_qoq[col + " Growth %"] = (
                            (combined_quarterly[col] - prev_values) /
                            prev_values.replace(0, pd.NA)
                        ) * 100
                    prev_values = combined_quarterly[col]

                # --- Format Growth % columns ---
                growth_cols = [c for c in quarterly_growth_qoq.columns if "Growth %" in c]
                for col in growth_cols:
                    quarterly_growth_qoq[col] = quarterly_growth_qoq[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # --- Reset index for Streamlit display ---
                quarterly_growth_qoq = quarterly_growth_qoq.reset_index()

               


                # --- Filter pivots by selected brands ---
                selected_brands = st.session_state.get("selected_brands", [])
                # pivot_monthly_filtered = pivot_monthly_clean_nonzero[
                #     pivot_monthly_clean_nonzero.index.get_level_values("Brand").isin(selected_brands) |
                #     pivot_monthly_clean_nonzero.index.get_level_values("Brand").str.contains("Total")
                # ]
                # pivot_monthly_growth_filtered = pivot_monthly_growth.loc[pivot_monthly_growth.index.isin(selected_brands)]
                # pivot_quarterly_filtered = pivot_quarterly.loc[pivot_quarterly.index.isin(selected_brands)]
                # quarterly_growth_filtered = quarterly_growth.loc[quarterly_growth.index.isin(selected_brands)]


                # --- Display Monthly Pivot ---
                st.subheader(f"📅 Monthly Data (NS) — {selected_seg}")
                st.dataframe(combined_pivot.round(0).astype(int))

                # -------------------------------------------------------------
                # 📝 Shared Comment Box for Monthly Pivot Table
                # -------------------------------------------------------------

                # Create session storage if not already exists
                if "monthly_pivot_comment" not in st.session_state:
                    st.session_state["monthly_pivot_comment"] = ""

                monthly_comment = st.text_area(
                    "📝 Comment on Monthly Data (NS):",
                    value=st.session_state["monthly_pivot_comment"],
                    key="monthly_pivot_comment_main"
                )

                # Update global storage
                st.session_state["monthly_pivot_comment"] = monthly_comment

                # --- Display Monthly vol Pivot ---
                st.subheader(f"📅 Monthly Data (Vol) — {selected_seg}")
                st.dataframe(combined_pivot_vol.round(0).astype(int))

                # If the index is a MultiIndex but has no names
                # if isinstance(pivot_monthly_growth.index, pd.MultiIndex):
                #     pivot_monthly_growth.index.names = ['Brand Family', 'Brand']

                # # --- Filter for selected brands ---
                # pivot_monthly_growth_filtered = monthly_growth_df.loc[
                #     monthly_growth_df.index.get_level_values("Brand").isin(selected_brands) |
                #     monthly_growth_df.index.get_level_values("Brand").str.contains("Total")
                # ]

                # Round numeric columns
                pivot_monthly_growth_filtered = monthly_growth_df.round(2)

                # Columns to exclude from formatting
                exclude_cols = ['Brand Family', 'Brand']

                # Determine columns to format (all except excluded)
                format_cols = [c for c in pivot_monthly_growth_filtered.columns if c not in exclude_cols]

                # Create a styled copy
                styled_growth = pivot_monthly_growth_filtered.copy()

                # Format numeric columns as percentages
                for col in format_cols:
                    styled_growth[col] = styled_growth[col].apply(
                        lambda x: f"{x:.2f}%" if pd.notnull(x) else ""
                    )
                # --- Display ---
                st.subheader(f"📈 JAJ Seg declines — {selected_seg}")
                st.dataframe(styled_growth)

                # # --- Display Quarterly Pivot ---
                # st.subheader(f"📆 Quarterly Totals — {selected_seg}")
                # st.dataframe(pivot_quarterly_filtered)

                # Streamlit Display
                st.subheader(f"📈 Quarterly NS & QoQ Growth — {selected_seg}")
                st.dataframe(quarterly_growth_qoq.round(0))


                # --- Annual NS M INR ---
                pivot_ns = pd.pivot_table(
                    df_filtered,
                    index='State',
                    columns='PRI Year',
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                )

                # Segment total row
                seg_total = pivot_ns.sum(axis=0)
                pivot_ns.loc["Seg Total"] = seg_total

                # --- State Share (%) ---
                pivot_share = pivot_ns.div(seg_total.values, axis=1) * 100
                pivot_share = pivot_share.round(2)

                # --- Annual Growth % ---
                pivot_growth = pivot_ns.pct_change(axis=1) * 100
                pivot_growth = pivot_growth.round(2)
                pivot_growth = pivot_growth.iloc[:, 1:]  # remove first year (NaN growth)

                # --- Convert Share % to string with % ---
                pivot_share_fmt = pivot_share.copy()
                pivot_share_fmt = pivot_share_fmt.applymap(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # --- Convert Annual Growth % to string with % ---
                pivot_growth_fmt = pivot_growth.copy()
                pivot_growth_fmt = pivot_growth_fmt.applymap(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # --- Quarterly NS M INR ---
                pivot_qtr = pd.pivot_table(
                    df_filtered,
                    index='State',
                    columns=['PRI Year', 'Quarter'],
                    values='NS M INR',
                    aggfunc='sum',
                    fill_value=0
                )

                # # --- Quarterly Growth % ---
                # qtr_growth = pivot_qtr.copy()
                # for year in pivot_qtr.columns.levels[0]:
                #     for q in ['Q1','Q2', 'Q3', 'Q4']:
                #         try:
                #             prev_q = f'Q{int(q[1])-1}'
                #             qtr_growth[(year, f'{q} Growth %')] = (
                #                 (pivot_qtr[(year, q)] - pivot_qtr[(year, prev_q)]) /
                #                 pivot_qtr[(year, prev_q)].replace(0, pd.NA)
                #             ) * 100
                #         except KeyError:
                #             continue
                # --- Flatten pivot for easier calculation ---
                pivot_qtr_flat = pivot_qtr.copy()
                pivot_qtr_flat.columns = [f"{year}_{q}" for year, q in pivot_qtr.columns]

                # Sort columns chronologically by year then quarter order
                years = sorted({col.split("_")[0] for col in pivot_qtr_flat.columns})
                quarter_order = ['Q1','Q2','Q3','Q4']
                cols_sorted = []
                for year in years:
                    for q in quarter_order:
                        col_name = f"{year}_{q}"
                        if col_name in pivot_qtr_flat.columns:
                            cols_sorted.append(col_name)
                pivot_qtr_flat = pivot_qtr_flat[cols_sorted]
                # st.dataframe(pivot_qtr_flat)

                # --- Compute QoQ growth across all quarters ---
                qtr_growth_flat = pivot_qtr_flat.copy()
                prev_values = None
                for col in pivot_qtr_flat.columns:
                    if prev_values is None:
                        qtr_growth_flat[col + " Growth %"] = pd.NA  # first quarter has no previous
                    else:
                        qtr_growth_flat[col + " Growth %"] = ((pivot_qtr_flat[col] - prev_values) / prev_values) * 100
                    prev_values = pivot_qtr_flat[col]

                # --- Rename quarters using custom map ---
                for col in qtr_growth_flat.columns:
                    # Only replace Q1-Q4 in the column part
                    for q, custom in quarter_custom_map.items():
                        if q in col:
                            new_col = col.replace(q, custom)
                            qtr_growth_flat.rename(columns={col: new_col}, inplace=True)

                # --- Optional: format Growth % as string with '%' for Streamlit ---
                growth_cols = [c for c in qtr_growth_flat.columns if "Growth %" in c]
                for col in growth_cols:
                    qtr_growth_flat[col] = qtr_growth_flat[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # # --- Display in Streamlit ---
                # st.subheader(f"📈 Quarterly NS & QoQ Growth — {selected_seg}")
                # keep only growth columns for display
                qtr_growth_flat = qtr_growth_flat[[c for c in qtr_growth_flat.columns if "Growth %" in c]]
                # st.dataframe(qtr_growth_flat)

                # --- Compute Quarterly Growth % with custom quarter names ---
                qtr_growth = pivot_qtr.copy()
                for year in pivot_qtr.columns.levels[0]:
                    for q in ['Q1','Q2','Q3','Q4']:
                        try:
                            prev_q = f'Q{int(q[1])-1}'  # previous quarter
                            # Compute growth %
                            qtr_growth[(year, f"{quarter_custom_map[q]} Growth %")] = (
                                (pivot_qtr[(year, q)] - pivot_qtr[(year, prev_q)]) /
                                pivot_qtr[(year, prev_q)].replace(0, pd.NA)
                            ) * 100
                        except KeyError:
                            continue
                        

                # --- Function to rename quarters in MultiIndex ---
                def rename_quarters(multiindex_cols, custom_map):
                    new_cols = []
                    for year, quarter in multiindex_cols:
                        # Replace quarter if in the map, else keep original
                        new_quarter = custom_map.get(quarter, quarter)
                        new_cols.append((year, new_quarter))
                    return pd.MultiIndex.from_tuples(new_cols)

                # --- Apply to quarterly NS pivot ---
                pivot_qtr.columns = rename_quarters(pivot_qtr.columns, quarter_custom_map)

                # --- Apply to quarterly growth pivot ---
                qtr_growth.columns = rename_quarters(qtr_growth.columns, quarter_custom_map)

                # # Flatten quarterly growth for concatenation
                # qtr_growth_flat = pd.DataFrame()
                # for year in qtr_growth.columns.levels[0]:
                #     for col in qtr_growth[year].columns:
                #         qtr_growth_flat[f"{year}_{col}"] = qtr_growth[(year, col)]


                # --- Flatten annual NS M, Share, Growth ---
                pivot_ns = pivot_ns.round(0).astype(int)
                pivot_ns_flat = pivot_ns.copy()
                pivot_ns_flat.columns = [f"{col}_NS M INR" for col in pivot_ns.columns]
                pivot_share_flat = pivot_share_fmt.copy()
                pivot_share_flat.columns = [f"{col}_Share %" for col in pivot_share_fmt.columns]
                pivot_growth_flat = pivot_growth_fmt.copy()
                pivot_growth_flat.columns = [f"{col}_Growth %" for col in pivot_growth_fmt.columns]

                # st.dataframe(pivot_ns_flat)
                # st.dataframe(pivot_share_flat)
                # st.dataframe(pivot_growth_flat)
                # st.dataframe(qtr_growth_flat)


                # --- Combine all in desired order ---
                final_df_state = pd.concat(
                    [pivot_ns_flat, pivot_share_flat, pivot_growth_flat, qtr_growth_flat],
                    axis=1
                )

                # --- Display ---
                st.subheader(f"📊 State Salience — {selected_seg}")
                st.dataframe(final_df_state)

                

                # --- Annual NS M INR ---
                pivot_ns = pd.pivot_table(
                    df_filtered,
                    index="State",
                    columns="PRI Year",
                    values="NS M INR",
                    aggfunc="sum",
                    fill_value=0
                )

                # df_filtered_defence = df_filtered[df_filtered["State"]=="Defence"]
                # st.write(f"Total DEFENCE NS M INR: {df_filtered_defence['NS M INR'].sum():,.0f} ₹ Mn")
                # st.write(df_filtered["State"].unique())


                # --- Segment total row ---
                seg_total = pivot_ns.sum(axis=0)
                pivot_ns.loc["Seg Total"] = seg_total

                # --- State Share (%) ---
                pivot_share = pivot_ns.div(seg_total.values, axis=1) * 100
                pivot_share = pivot_share.round(2)

                # --- Annual Growth % ---
                pivot_growth = pivot_ns.pct_change(axis=1) * 100
                pivot_growth = pivot_growth.round(2)

                pivot_growth = pivot_growth.iloc[:, 1:]  # remove first year (NaN growth)

                # --- Convert Share % to string with % ---
                pivot_share_fmt = pivot_share.copy()
                pivot_share_fmt = pivot_share_fmt.applymap(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # --- Convert Annual Growth % to string with % ---
                pivot_growth_fmt = pivot_growth.copy()
                pivot_growth_fmt = pivot_growth_fmt.applymap(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

                # --- Quarterly NS M INR ---
                pivot_qtr = pd.pivot_table(
                    df_filtered,
                    index="State",
                    columns=["PRI Year", "Quarter"],
                    values="NS M INR",
                    aggfunc="sum",
                    fill_value=0
                )

                # --- Quarterly Growth % ---
                qtr_growth = pivot_qtr.copy()
                for year in pivot_qtr.columns.get_level_values(0).unique():
                    for q, prev in [("Q2", "Q1"), ("Q3", "Q2"), ("Q4", "Q3")]:
                        if (year, q) in pivot_qtr.columns and (year, prev) in pivot_qtr.columns:
                            base = pivot_qtr[(year, prev)].replace(0, np.nan)
                            qtr_growth[(year, f"{q} Growth %")] = (pivot_qtr[(year, q)] - pivot_qtr[(year, prev)]) / base * 100

                # --- Flatten quarterly growth for concatenation ---
                # qtr_growth_flat = pd.DataFrame()
                # for year in qtr_growth.columns.get_level_values(0).unique():
                #     for col in qtr_growth[year].columns:
                #         qtr_growth_flat[f"{year}_{col}"] = qtr_growth[(year, col)]

                # --- Flatten annual NS M, Share, Growth ---
                pivot_ns = pivot_ns.round(0).astype(int)
                pivot_ns_flat = pivot_ns.copy()
                pivot_ns_flat.columns = [f"{col}_NS M INR" for col in pivot_ns.columns]
                pivot_share_flat = pivot_share_fmt.copy()
                pivot_share_flat.columns = [f"{col}_Share %" for col in pivot_share_fmt.columns]
                pivot_growth_flat = pivot_growth_fmt.copy()
                pivot_growth_flat.columns = [f"{col}_Growth %" for col in pivot_growth_fmt.columns]

                # --- Combine all in desired order ---
                final_df = pd.concat(
                    [pivot_ns_flat, pivot_share_flat, pivot_growth_flat, qtr_growth_flat],
                    axis=1
                )

                # --- monthly NS M INR ---
                pivot_monthly = pd.pivot_table(
                    df_filtered,
                    index="State",
                    columns=["PRI Year", "Month"],
                    values="NS M INR",
                    aggfunc="sum",
                    fill_value=0
                )

         

                # ==============================
                # 🔽 Interactive Top-States Filter
                # ==============================
                st.write("## State Selection Based on Market Contribution")

                # --- Step 1: Select granularity ---
                granularity = st.radio(
                    "Select level for filtering:",
                    ["Yearly", "Quarterly", "Monthly"],
                    horizontal=True
                )

                # --- Step 2: Select year, quarter, or month ---
                if granularity == "Yearly":
                    available_years = sorted(pivot_ns.columns)
                    selected_year = st.selectbox("Select Year:", available_years)
                    sales_df = pivot_ns[selected_year]

                elif granularity == "Quarterly":
                    # --- Custom mapping for quarter display ---
                    quarter_custom_map = {
                        'Q1': 'JAS',   # July–Aug–Sep
                        'Q2': 'OND',   # Oct–Nov–Dec
                        'Q3': 'JFM',   # Jan–Feb–Mar
                        'Q4': 'AMJ'    # Apr–May–Jun
                    }

                    available_years = sorted({y for y, _ in pivot_qtr.columns})
                    selected_year = st.selectbox("Select Year:", available_years)

                    # Extract available quarters for selected year
                    available_quarters = sorted({q for y, q in pivot_qtr.columns if y == selected_year})

                    # Build display labels for quarters using custom map
                    quarter_labels = [quarter_custom_map.get(q, q) for q in available_quarters]
                    label_to_quarter = {quarter_custom_map.get(q, q): q for q in available_quarters}

                    # Streamlit selectbox shows friendly label but maps back to actual quarter code
                    selected_quarter_label = st.selectbox("Select Quarter:", quarter_labels)
                    selected_quarter = label_to_quarter[selected_quarter_label]

                    # Get corresponding data
                    sales_df = pivot_qtr[(selected_year, selected_quarter)]


                elif granularity == "Monthly":
                    # --- Define fiscal month order: July → June ---
                    fiscal_month_order = [
                        "July", "August", "September", "October", "November", "December",
                        "January", "February", "March", "April", "May", "June"
                    ]

                    available_years = sorted({y for y, _ in pivot_monthly.columns})
                    selected_year = st.selectbox("Select Year:", available_years)

                    # --- Extract available months for selected year ---
                    available_months_raw = [m for y, m in pivot_monthly.columns if y == selected_year]

                    # --- Reorder months by fiscal order (July → June) ---
                    available_months = [m for m in fiscal_month_order if m in available_months_raw]

                    # --- Month selector in fiscal order ---
                    selected_month = st.selectbox("Select Month (Fiscal Order: Jul–Jun):", available_months)

                    # --- Get corresponding data ---
                    sales_df = pivot_monthly[(selected_year, selected_month)]


                # --- Step 3: Compute Market Share ---
                st.write(f"### Select States Contributing to Total Market ({granularity}: {selected_year})")

                # Remove "Seg Total" if exists
                latest_sales = sales_df.loc[sales_df.index != "Seg Total"].sort_values(ascending=False)
                total_sales = latest_sales.sum()
                share_pct = (latest_sales / (total_sales if total_sales != 0 else np.nan) * 100).round(1)

                # --- Step 4: Threshold slider ---
                cut = st.slider(
                    "Cumulative share threshold (%)",
                    min_value=50, max_value=100, value=85, step=1,
                    help="States up to this cumulative share (plus the next state) are eligible by default."
                )

                # --- Step 5: Calculate cumulative share and eligible states ---
                cumulative_share = (latest_sales.cumsum() / (total_sales if total_sales != 0 else np.nan)) * 100
                within = cumulative_share[cumulative_share <= cut]
                eligible = within.index.tolist()
                if len(within) < len(cumulative_share):
                    eligible.append(cumulative_share.index[len(within)])  # include next state

                # --- Step 6: Initialize session_state ---
                if "state_selected" not in st.session_state:
                    st.session_state.state_selected = {}

                for s in eligible:
                    st.session_state.state_selected.setdefault(s, True)

                # --- Step 7: Select All / Clear All ---
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Select all"):
                        for s in eligible:
                            st.session_state.state_selected[s] = True
                with c2:
                    if st.button("Clear all"):
                        for s in eligible:
                            st.session_state.state_selected[s] = False

                # --- Step 8: Display checkboxes with shares ---
                cols = st.columns(4)
                for i, s in enumerate(eligible):
                    col = cols[i % 4]
                    label = f"{s} — {share_pct.get(s, np.nan):.1f}%"
                    st.session_state.state_selected[s] = col.checkbox(
                        label=label,
                        value=st.session_state.state_selected[s],
                        key=f"cb_state_{granularity}_{s}"
                    )

                # --- Step 9: Get final selected states ---
                chosen_states = [s for s in latest_sales.index if s in eligible and st.session_state.state_selected.get(s, False)]

                # --- Step 10: Filter & display ---
                if chosen_states:
                    final_df_filtered = final_df.loc[chosen_states]
                    st.subheader(f"📊 Filtered State Salience ({granularity}: {selected_year})")
                    st.dataframe(final_df_filtered, use_container_width=True)
                else:
                    st.warning("Select at least one state to view.")


             



                df_filtered = df_filtered[df_filtered["State"].isin(chosen_states)]

                # --- Pivot Brand × Year ---
                pivot_brand = pd.pivot_table(
                    df_filtered,
                    index="State",
                    columns=["Brand", "PRI Year"],
                    values="NS M INR",
                    aggfunc="sum",
                    fill_value=0,
                ).sort_index(axis=1, level=[0, 1])

                # --- Segment Totals (All Brands in Segment) ---
                pivot_segment = pivot_brand.groupby(level=1, axis=1).sum()
                pivot_segment.columns = [f"Segment Total {y}" for y in pivot_segment.columns]

                # --- Identify available years ---
                all_years = sorted(df_filtered["PRI Year"].dropna().unique())

                # --- Calculate Segment Share (%) ---
                seg_share_dict = {}
                for brand, year in pivot_brand.columns:
                    seg_total_col = f"Segment Total {year}"
                    share = (
                        pivot_brand[(brand, year)] / pivot_segment[seg_total_col].replace(0, pd.NA)
                    ) * 100
                    seg_share_dict[(brand, f"Share {year} %")] = share.round(2)
                seg_share_df = pd.DataFrame(seg_share_dict, index=pivot_brand.index)

                # --- Identify Brands ≥ 1% in Any Year ---
                brands_to_keep = []
                for brand in pivot_brand.columns.get_level_values(0).unique():
                    brand_shares = seg_share_df[[c for c in seg_share_df.columns if c[0] == brand]]
                    if (brand_shares >= 1).any().any():
                        brands_to_keep.append(brand)

                if not brands_to_keep:
                    st.warning("⚠️ No brands found with ≥ 1% segment share in any year.")
                else:
                    # --- Keep Only Those Brands ---
                    pivot_brand = pivot_brand[brands_to_keep]

                    # --- Calculate Brand Growth ---
                    # growth_dict = {}
                    # for brand in brands_to_keep:
                    #     for i in range(1, len(all_years)):
                    #         y1, y2 = all_years[i - 1], all_years[i]
                    #         if (brand, y1) in pivot_brand.columns and (brand, y2) in pivot_brand.columns:
                    #             v1 = pivot_brand[(brand, y1)]
                    #             v2 = pivot_brand[(brand, y2)]
                    #             growth = ((v2 - v1) / v1.replace(0, pd.NA)) * 100
                    #             growth_dict[(brand, f"Growth {y1}-{y2} %")] = growth.round(2)

                    # --- Calculate Brand Growth (safe) ---
                    growth_dict = {}
                    for brand in brands_to_keep:
                        for i in range(1, len(all_years)):
                            y1, y2 = all_years[i - 1], all_years[i]
                            try:
                                v1 = pivot_brand[(brand, y1)]
                                v2 = pivot_brand[(brand, y2)]
                                growth = ((v2 - v1) / v1.replace(0, pd.NA)) * 100
                                growth_dict[(brand, f"Growth {y1}-{y2} %")] = growth.round(2)
                            except KeyError:
                                # Skip brand-year combos not present
                                continue
                    growth_df = pd.DataFrame(growth_dict, index=pivot_brand.index)

                    # --- Calculate Segment Growth ---
                    seg_growth_dict = {}
                    for i in range(1, len(all_years)):
                        y1, y2 = all_years[i - 1], all_years[i]
                        c1, c2 = f"Segment Total {y1}", f"Segment Total {y2}"
                        seg_growth = ((pivot_segment[c2] - pivot_segment[c1]) / pivot_segment[c1].replace(0, pd.NA)) * 100
                        seg_growth_dict[f"Segment Growth {y1}-{y2} %"] = seg_growth.round(2)
                    seg_growth_df = pd.DataFrame(seg_growth_dict, index=pivot_brand.index)

                    # --- Calculate Relative Growth (Brand - Segment) ---
                    relative_growth_dict = {}
                    for brand in brands_to_keep:
                        for i in range(1, len(all_years)):
                            y1, y2 = all_years[i - 1], all_years[i]
                            brand_col = (brand, f"Growth {y1}-{y2} %")
                            seg_col = f"Segment Growth {y1}-{y2} %"
                            if brand_col in growth_df.columns and seg_col in seg_growth_df.columns:
                                # relative = growth_df[brand_col] - seg_growth_df[seg_col]
                                # relative_growth_dict[(brand, f"Relative Growth {y1}-{y2} %")] = relative.round(2)
                                try:
                                    relative = growth_df[brand_col] - seg_growth_df[seg_col]
                                    relative_growth_dict[(brand, f"Diff Growth {y1}-{y2} %")] = relative.round(2)
                                except KeyError:
                                    continue
                    relative_growth_df = pd.DataFrame(relative_growth_dict, index=pivot_brand.index)

                    # --- Determine the latest period automatically ---
                    if len(all_years) < 2:
                        st.warning("⚠️ Not enough years available to compute growth (need at least 2).")
                        st.stop()

                    # Get the latest two years (e.g., A24, A25)
                    y1, y2 = all_years[-2], all_years[-1]
                    period = f"{y1}-{y2}"
                    # st.info(f"📆 Automatically showing results for the latest period: {period}")

                    # --- Get brands present in pivot columns ---
                    available_brands = pivot_brand.columns.get_level_values(0).unique()
                    selected_brands = st.session_state.get("selected_brands", [])

                    # --- Only keep brands that exist in data ---
                    brands_to_display = [b for b in selected_brands if b in available_brands]

                    if not brands_to_display:
                        st.warning("⚠️ None of the selected brands are available in the data for the selected filters.")
                        st.stop()

                    # --- Now ensure only brands with data in BOTH y1 and y2 (based on pivot_brand) ---
                    brands_for_period = []
                    for b in brands_to_display:
                        has_y1 = (b, y1) in pivot_brand.columns
                        has_y2 = (b, y2) in pivot_brand.columns
                        if has_y1 and has_y2:
                            brands_for_period.append(b)

                    if not brands_for_period:
                        st.warning(f"⚠️ None of the selected brands have data for both {y1} and {y2}.")
                        st.stop()

                    # --- Filter brand-level data ---
                    idx = pd.IndexSlice
                    pivot_brand = pivot_brand.loc[:, idx[brands_for_period, :]]
                    seg_share_df = seg_share_df.loc[:, idx[brands_for_period, :]]
                    growth_df = growth_df.loc[:, idx[brands_for_period, :]]
                    relative_growth_df = relative_growth_df.loc[:, idx[brands_for_period, :]]

                    # --- Prepare segment growth column for the period ---
                    seg_col = f"Segment Growth {period} %"
                    if seg_col not in seg_growth_df.columns:
                        st.warning(f"⚠️ Segment growth column '{seg_col}' not found. Please check data consistency.")
                        st.stop()

                    seg_df = seg_growth_df[[seg_col]].copy()
                    seg_df.columns = pd.MultiIndex.from_tuples([("Segment", seg_col)])

                    # --- Combine Segment + Brand + Relative Growth for this period ---
                    combined_for_period = seg_df.copy()

                    for brand in brands_for_period:
                        b_col = f"Growth {period} %"
                        r_col = f"Diff Growth {period} %"

                        if (brand, b_col) in growth_df.columns and (brand, r_col) in relative_growth_df.columns:
                            combined_for_period[(brand, b_col)] = growth_df[(brand, b_col)]
                            combined_for_period[(brand, r_col)] = relative_growth_df[(brand, r_col)]

                    # --- MultiIndex columns (period > brand > metric) ---
                    combined_for_period.columns = pd.MultiIndex.from_tuples(
                        [(period, col[0], col[1]) for col in combined_for_period.columns]
                    )

                    # --- Display final table ---
                    final_display_df = combined_for_period.sort_index(axis=1, level=[0, 1, 2])

                    # --- Identify columns containing 'Segment' ---
                    segment_cols = [col for col in final_display_df.columns if "Segment" in col]

                    # --- Get all remaining columns ---
                    other_cols = [col for col in final_display_df.columns if col not in segment_cols]

                    # --- Reorder so Segment columns come first ---
                    final_display_df = final_display_df[segment_cols + other_cols]

                    # --- Display ---
                    # st.dataframe(final_display_df.style.format("{:.2f}%"))
                    # --- Step 1: Remove the first (year) level from MultiIndex columns ---
                    df = final_display_df.copy()
                    df.columns = df.columns.droplevel(0)  # drop 'A24-A25'

                    # --- Step 2: Identify segment and brand columns ---
                    segment_cols = [col for col in df.columns if col[0] == "Segment"]
                    brand_cols = [col for col in df.columns if col[0] != "Segment"]

                    # --- Step 3: Sort brand columns (Growth first, then Diff) ---
                    ordered_brand_cols = []
                    brands = sorted(set([b for b, _ in brand_cols]))

                    for b in brands:
                        growth_col = (b, f"Growth {period} %")
                        diff_col   = (b, f"Diff Growth {period} %")
                        if growth_col in df.columns:
                            ordered_brand_cols.append(growth_col)
                        if diff_col in df.columns:
                            ordered_brand_cols.append(diff_col)

                    # --- Step 4: Combine Segment first, then each brand pair ---
                    final_display_df = df[segment_cols + ordered_brand_cols]

                    # --- Step 5: Optional - Clean metric names (remove redundant year text) ---
                    final_display_df.columns = pd.MultiIndex.from_tuples([
                        (b, m.replace(f" {period}", "").strip()) for b, m in final_display_df.columns
                    ])
                    

                    # --- 1️⃣ Get Segment columns first ---
                    segment_cols = [col for col in final_display_df.columns if 'Segment' in col]

                    # --- 2️⃣ Build ordered list of brand columns ---
                    ordered_cols = []
                    for b in brand_order:
                        # Match columns belonging to this brand
                        growth_cols = [col for col in final_display_df.columns if col[0] == b and 'Growth' in col[1] and 'Diff' not in col[1]]
                        diff_cols   = [col for col in final_display_df.columns if col[0] == b and 'Diff' in col[1]]
                        
                        # Append growth first, then diff
                        ordered_cols.extend(growth_cols + diff_cols)

                    # --- 3️⃣ Rebuild final DataFrame with segment first, then sorted brands ---
                    final_display_df = final_display_df[segment_cols + ordered_cols]

                    # --- Step 5.5: Rename "Diff Growth %" to "BTM" for all brands ---
                    final_display_df.columns = pd.MultiIndex.from_tuples([
                        (b, m.replace("Diff Growth %", "BTM")) for b, m in final_display_df.columns
                    ])


                    # --- Step 6: Display ---
                    st.subheader(f"📊 Segment, Brand & Growth Comparison — {selected_seg}")
                    st.dataframe(final_display_df.style.format("{:.2f}%"))

                    # -------------------------------------------------------------
                    # 📝 Shared Comment Box for Segment/Brand/Growth Table
                    # -------------------------------------------------------------

                    # Create session storage if not already created
                    if "seg_brand_growth_comment" not in st.session_state:
                        st.session_state["seg_brand_growth_comment"] = ""

                    # Text area linked to session_state
                    seg_brand_comment = st.text_area(
                        "📝 Comment on Segment, Brand & Growth Comparison:",
                        value=st.session_state["seg_brand_growth_comment"],
                        key="seg_brand_growth_comment_box"
                    )

                    # Update session state
                    st.session_state["seg_brand_growth_comment"] = seg_brand_comment





                    # st.subheader(f"📊 Segment, Brand & Relative Growth Comparison — {selected_seg}")
                    # st.dataframe(final_display_df.style.format("{:.2f}%"))

            

                    # --- Combine only Brand Growth + Segment Growth ---
                    # final_df = pd.concat([growth_df, seg_growth_df], axis=1)
                    final_df_growth = pd.concat([seg_growth_df, growth_df], axis=1)





                import pandas as pd
                import numpy as np


                st.subheader(f"📊Zonal View — {selected_seg}")
                # # Brand picker
                # brand_opts = sorted(df_filtered_zone["Brand"].dropna().unique().tolist())  # options
                # selected_brands = st.multiselect(
                #     "Include Brands",
                #     options=brand_opts,
                #     default=brand_opts,
                #     key="brand_multiselect",
                #     placeholder="Choose brands…",
                # )

                selected_brands = st.session_state.get("selected_brands", [])
                # df_filtered_zone = df_filtered_zone[df_filtered_zone["State"]!='Delhi']


                def build_zone_brand_matrix(df_filtered_zone: pd.DataFrame,
                                            selected_years: list[str],
                                            zones_order: list[str] | None = None) -> pd.DataFrame:
                    # 0) Guard inputs
                    req = {"Brand", "Zone", "PRI Year", "NS M INR"}
                    missing = req - set(df_filtered_zone.columns)
                    if missing:
                        raise ValueError(f"Missing columns: {missing}")

                    # 1) Base pivot: Brand × (Zone, Year) totals
                    base = pd.pivot_table(
                        df_filtered_zone,
                        index="Brand",
                        columns=["Zone", "PRI Year"],
                        values="NS M INR",
                        aggfunc="sum",
                        fill_value=0,
                    )

                    # Keep only selected years present
                    present_years = base.columns.get_level_values(1).unique().tolist()
                    years = [y for y in selected_years if y in present_years]
                    if not years:
                        raise ValueError("None of the selected years exist in the data.")

                    # Zone ordering
                    zones = zones_order or list(base.columns.get_level_values(0).unique())

                    # Helpers
                    def yoy_series(new_s: pd.Series, old_s: pd.Series) -> pd.Series:
                        denom = old_s.replace(0, np.nan)
                        return (new_s - old_s) / denom * 100.0

                    def safe_yoy(new_v, old_v):
                        if pd.isna(old_v) or (isinstance(old_v, (int, float, np.floating)) and old_v == 0):
                            return np.nan
                        return (new_v - old_v) / old_v * 100.0

                    # 2) Zone totals across brands (for YoY reference and Zone Total row)
                    zone_year_tot = base.sum(axis=0)  # Series indexed by (Zone, Year)

                    # Per-zone YoY map (scalars)
                    zone_yoy = {}
                    for z in zones:
                        zmap = {}
                        if "A24" in years and "A23" in years and (z, "A24") in zone_year_tot and (z, "A23") in zone_year_tot:
                            zmap["A24"] = safe_yoy(zone_year_tot[(z, "A24")], zone_year_tot[(z, "A23")])
                        if "A25" in years and "A24" in years and (z, "A25") in zone_year_tot and (z, "A24") in zone_year_tot:
                            zmap["A25"] = safe_yoy(zone_year_tot[(z, "A25")], zone_year_tot[(z, "A24")])
                        zone_yoy[z] = zmap

                    # 3) Overall (all-zone) totals and YoY for the "All Zones Total" top row (scalars)
                    overall_year_tot = (
                        df_filtered_zone.groupby("PRI Year")["NS M INR"].sum().reindex(years).fillna(0)
                    )
                    overall_yoy = {}
                    if "A24" in years and "A23" in years:
                        overall_yoy["A24"] = safe_yoy(overall_year_tot.get("A24", np.nan), overall_year_tot.get("A23", np.nan))
                    if "A25" in years and "A24" in years:
                        overall_yoy["A25"] = safe_yoy(overall_year_tot.get("A25", np.nan), overall_year_tot.get("A24", np.nan))

                    # 4) Build per-zone blocks
                    zone_blocks = []
                    for z in zones:
                        cols = {}
                        # Year Totals per brand
                        for y in years:
                            if (z, y) in base.columns:
                                cols[f"{y} Total"] = base[(z, y)]
                            else:
                                cols[f"{y} Total"] = pd.Series(index=base.index, dtype=float)
                        # Growths per brand (Series)
                        if "A24" in years and "A23" in years and (z, "A24") in base.columns and (z, "A23") in base.columns:
                            cols["A24 Gr"] = yoy_series(base[(z, "A24")], base[(z, "A23")]).round(2)
                        if "A25" in years and "A24" in years and (z, "A25") in base.columns and (z, "A24") in base.columns:
                            brand_a25 = yoy_series(base[(z, "A25")], base[(z, "A24")]).round(2)
                            cols["A25 Gr"] = brand_a25
                            # BTM = Brand A25 YoY − Zone A25 YoY (scalar applied row-wise)
                            z_ref = zone_yoy.get(z, {}).get("A25", np.nan)
                            cols["BTM"] = (brand_a25 - z_ref).round(2)
                        block = pd.DataFrame(cols, index=base.index)
                        block.columns = pd.MultiIndex.from_product([[z], block.columns])
                        zone_blocks.append(block)

                    # 5) Combine blocks horizontally
                    final = pd.concat(zone_blocks, axis=1)

                    # 6) Summary rows on top: "Grand Total" (overall repeated) and "Zone Total" (per-zone sums)
                    zone_total_row = {}
                    for z in zones:
                        for y in years:
                            zone_total_row[(z, f"{y} Total")] = zone_year_tot.get((z, y), np.nan)
                        if (z, "A24 Gr") in final.columns:
                            a24 = zone_year_tot.get((z, "A24"), np.nan)
                            a23 = zone_year_tot.get((z, "A23"), np.nan)
                            zone_total_row[(z, "A24 Gr")] = safe_yoy(a24, a23)
                        if (z, "A25 Gr") in final.columns:
                            a25 = zone_year_tot.get((z, "A25"), np.nan)
                            a24 = zone_year_tot.get((z, "A24"), np.nan)
                            zone_total_row[(z, "A25 Gr")] = safe_yoy(a25, a24)
                        if (z, "BTM") in final.columns:
                            zone_total_row[(z, "BTM")] = np.nan  # optional blank

                    all_zones_row = {}
                    for z in zones:
                        for y in years:
                            all_zones_row[(z, f"{y} Total")] = overall_year_tot.get(y, np.nan)
                        if (z, "A24 Gr") in final.columns:
                            all_zones_row[(z, "A24 Gr")] = overall_yoy.get("A24", np.nan)
                        if (z, "A25 Gr") in final.columns:
                            all_zones_row[(z, "A25 Gr")] = overall_yoy.get("A25", np.nan)
                        if (z, "BTM") in final.columns:
                            z_ref = zone_yoy.get(z, {}).get("A25", np.nan)
                            all_zones_row[(z, "BTM")] = (overall_yoy.get("A25", np.nan) - z_ref) if pd.notna(z_ref) else np.nan

                    top_rows = pd.DataFrame.from_dict(
                        {"Grand Total": all_zones_row, "Zone Total": zone_total_row}, orient="index"
                    ).reindex(columns=final.columns)

                    final = pd.concat([top_rows, final], axis=0).round(2)
                    final.index.name = None
                    return final
                

                selected_years = ["A23", "A24", "A25"]
                zones_order = ["East Zone", "North Zone", "South Zone", "West+CSD Zone"]

                # Guard when nothing selected
                if not selected_brands:
                    st.warning("Select at least one brand to build the table.")
                else:
                    df_brand = df_filtered_zone[df_filtered_zone["Brand"].isin(selected_brands)]
                    final_df_zone = build_zone_brand_matrix(df_brand, selected_years, zones_order)
                    #Display

                    percent_name_keys = ("Gr", "Gr %", "BTM", "BTM %")

                    def percentify_df(df):
                        # Works for both single and MultiIndex columns
                        cols = df.columns
                        for c in cols:
                            name = c[1] if isinstance(c, tuple) and len(c) > 1 else c
                            if any(key in str(name) for key in percent_name_keys):
                                df[c] = df[c].apply(lambda val: "" if pd.isnull(val) or np.isnan(val) else f"{val:.2f}%")
                        return df

                    styled_df = percentify_df(final_df_zone.copy())
                    st.dataframe(styled_df)


                    # st.dataframe(final_df_zone.style.format("{:.2f}"))


                from io import BytesIO

                def export_dfs_to_excel(
                    sheets: list[tuple[pd.DataFrame, str]],
                    index=True,
                    percent_name_keys=("Gr", "Gr %", "BTM", "BTM %"),
                    value_num_format="#,##0",
                    percent_num_format="0.0%",
                    remove_multiindex_blank_row=True,  # NEW
                ) -> BytesIO:
                    """
                    sheets: list of (df, sheet_name)
                    index: include index in Excel
                    percent_name_keys: column name substrings that indicate percent columns
                    Returns an in-memory BytesIO of the xlsx.
                    """
                    # Keep per-sheet header row counts so we can fix after write
                    header_rows_map = {name: (df.columns.nlevels if isinstance(df.columns, pd.MultiIndex) else 1)
                                    for df, name in sheets}


                    buf = BytesIO()
                    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:  # XlsxWriter supports column formats well
                        for df, sheet_name in sheets:
                            # Work on a copy for Excel-only transforms
                            df_x = df.copy()

                            # Detect percent vs value columns for flat or MultiIndex columns
                            if isinstance(df_x.columns, pd.MultiIndex):
                                metrics = df_x.columns.get_level_values(-1).astype(str)
                                is_pct = metrics.str.endswith("Gr") | metrics.str.endswith("Gr %") | (metrics == "BTM") | (metrics == "BTM %")
                                pct_cols = list(df_x.columns[is_pct])
                                val_cols = list(df_x.columns[~is_pct])
                            else:
                                names = pd.Index(df_x.columns).astype(str)
                                is_pct = (
                                    names.str.endswith("Gr") | names.str.endswith("Gr %") |
                                    (names == "BTM") | (names == "BTM %")
                                )
                                pct_cols = list(df_x.columns[is_pct])
                                val_cols = list(df_x.columns[~is_pct])

                            # Ensure numeric types where possible
                            for c in pct_cols + val_cols:
                                # safe numeric coercion without breaking non-numeric columns
                                if pd.api.types.is_numeric_dtype(df_x[c]):
                                    continue
                                try:
                                    df_x[c] = pd.to_numeric(df_x[c], errors="ignore")
                                except Exception:
                                    pass

                            # Percent columns: scale by 0.01 so Excel '0.0%' is correct
                            # Only scale numeric series
                            for c in pct_cols:
                                if pd.api.types.is_numeric_dtype(df_x[c]):
                                    df_x[c] = df_x[c] / 100.0

                            # Write the frame
                            df_x.to_excel(writer, sheet_name=sheet_name, index=index)

                            # Apply formats
                            wb = writer.book
                            ws = writer.sheets[sheet_name]
                            fmt_val = wb.add_format({"num_format": value_num_format})      # integer with thousands
                            fmt_pct = wb.add_format({"num_format": percent_num_format})    # one-decimal percent

                            # Column positions in Excel (account for index column)
                            start_col = 1 if index else 0
                            # Build a map from DataFrame column order to Excel column index
                            excel_col_idx = {col: (start_col + i) for i, col in enumerate(df_x.columns)}

                            # Set formats per column
                            for c in val_cols:
                                # Only set format if column is numeric
                                if pd.api.types.is_numeric_dtype(df_x[c]):
                                    col_ix = excel_col_idx[c]
                                    ws.set_column(col_ix, col_ix, None, fmt_val)
                            for c in pct_cols:
                                if pd.api.types.is_numeric_dtype(df_x[c]):
                                    col_ix = excel_col_idx[c]
                                    ws.set_column(col_ix, col_ix, None, fmt_pct)

                            # Optional: freeze header + index row, and add autofilter
                            header_rows = df_x.columns.nlevels if isinstance(df_x.columns, pd.MultiIndex) else 1
                            ws.freeze_panes(header_rows, start_col)  # freeze header rows and index col
                            # Add autofilter over header range
                            nrows, ncols = df_x.shape
                            #ws.autofilter(0, 0, header_rows + nrows - 1, start_col + ncols - 1)

                    buf.seek(0)
                    if remove_multiindex_blank_row:
                        wb = load_workbook(buf)
                        for df, sheet_name in sheets:
                            if isinstance(df.columns, pd.MultiIndex):
                                ws = wb[sheet_name]
                                header_rows = header_rows_map[sheet_name]  # e.g., 2 for Zone/Metric
                                ws.delete_rows(header_rows + 1)           # remove blank row (row 3)  
                        out = BytesIO()
                        wb.save(out)
                        out.seek(0)
                        return out
                    return buf



                # Example DataFrames you already computed on the page
                # final_df_zone, final_df_states, final_df_brands = ...

                sheets_to_write = [
                    (final_mfs_review, "Manufacturing Co. View"),
                    # (pivot_mfs_growth_fmt, "Manufacturing Co. Growth"),
                    (export_display_df, "Brand View"),
                    (combined_pivot,   "NS Monthly Pivot"),
                    (styled_growth, "JAJ Seg declines"),
                    (quarterly_growth_qoq, "Quarterly NS & QoQ Growth"),
                    # (quarterly_growth, "Quarterly Growth Rates (%)"),
                    (final_df_state, "State Salience"),
                    (final_df_filtered, "Filtered State Salience"),
                    (final_display_df, "Brand State View"),
                    # (final_df_growth, "Growth Rate Analysis"),
                    (final_df_zone, "Zonal View"),
                ]

                xlsx_bytes = export_dfs_to_excel(sheets_to_write, index=True)

                st.download_button(
                    "Download all sheets",
                    data=xlsx_bytes,
                    file_name="BTM_Analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )



        except ValueError:
            st.error("❌ The sheet named 'Data' was not found in the uploaded Excel file.")
        except Exception as e:
            st.error(f"⚠️ Error loading file: {e}")
    else:
        st.info("Please upload an Excel file to continue.")

with tab2:

    # st.title("📊 PRI Maestria Data Tool")
    st.subheader("Upload Input Files")

    # -----------------------------
    # Session state init
    # -----------------------------
    for key, default in {
        "submitted": False,
        "selected_sheet": None,
        "prev_file": None,
        "df": None,
        "dims": []
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    # -----------------------------
    # File upload + sheet selection (gated)
    # -----------------------------
    uploaded_file = st.sidebar.file_uploader("Upload the Excel Data File", type=["xlsx"], key="excel_file")

    # Reset state if file changes
    cur_file = getattr(uploaded_file, "name", None)
    if cur_file != st.session_state.prev_file:
        st.session_state.submitted = False
        st.session_state.selected_sheet = None
        st.session_state.prev_file = cur_file
        st.session_state.df = None  # reset cached df too

    if uploaded_file is None:
        st.info("Upload a file to begin.")
        st.stop()

    xls = pd.ExcelFile(uploaded_file)
    with st.form("sheet_picker_form", clear_on_submit=False):
        sheet = st.selectbox(
            "Select a sheet",
            options=xls.sheet_names,
            index=None,
            placeholder="Choose a sheet…",
            key="sheet_name"
        )
        clicked = st.form_submit_button("Load sheet", type="primary")

    if clicked and sheet is not None:
        st.session_state.submitted = True
        st.session_state.selected_sheet = sheet

    if not (st.session_state.submitted and st.session_state.selected_sheet):
        st.info("Select a sheet and click Load sheet to proceed.")
        st.stop()

    # Load df once per file/sheet
    if st.session_state.df is None:
        df = pd.read_excel(uploaded_file, sheet_name=st.session_state.selected_sheet)
        st.session_state.df = df.copy()
    else:
        df = st.session_state.df

    st.success(f"Loaded sheet: {st.session_state.selected_sheet}")
    # Optional: show raw df preview (comment out if not needed)
    st.dataframe(df, use_container_width=True)

    # -----------------------------
    # Segment selection
    # -----------------------------
    # Build segment list safely
    if "vjp_occasion_category" not in df.columns:
        st.error("Column 'vjp_occasion_category' is missing in this sheet.")
        st.stop()

    segment_options = (
        df["vjp_occasion_category"].dropna().unique().tolist()
    )
    segment_options = sorted(segment_options)

    st.header("Select segments:")
    selected_seg = st.multiselect(
        "Select Revised Segment(s):",
        options=segment_options,
        default=[]
        # default=segment_options[:1],  # optional, default selection
    )

    if not selected_seg:
        st.warning("Please select at least one segment to continue.")
        st.stop()

    # filtered_df = df[df['vjp_occasion_category'].isin(selected_seg)]

    # 2. Specify custom name for selected segments
    default_name = selected_seg[0] if len(selected_seg) == 1 else " + ".join(selected_seg)
    custom_name = st.text_input(
        "Enter a new label for the segment(s):",
        value=default_name
    )

    # 3. Rename selected segments in the full DataFrame
    df_renamed = df.copy()
    df_renamed.loc[
        df_renamed["vjp_occasion_category"].isin(selected_seg), "vjp_occasion_category"
    ] = custom_name

    filtered_df = df_renamed[df_renamed["vjp_occasion_category"] == custom_name]

    st.header("Filtered Data:")
    st.dataframe(filtered_df, use_container_width=True)

    # -----------------------------
    # Dimension picking (single-add) + run
    # -----------------------------
    # Guard for empty filtered_df
    if filtered_df.empty:
        st.warning("No rows for the selected segment.")
        st.stop()

    # Anchor column must exist
    if "QAge" not in filtered_df.columns:
        st.error("Anchor column 'QAge' is missing.")
        st.stop()

    anchor = filtered_df.columns.get_loc("QAge")
    dim_candidates = [c for c in filtered_df.columns[:anchor+1] if c != "vjp_occasion_category"]

    # Single add UI
    st.header("Add grouping dimensions:")
    # dim_to_add = st.selectbox(
    #     "Add a grouping dimension",
    #     options=[d for d in dim_candidates if d not in st.session_state.dims],
    #     index=None
    # )
    # add_clicked = st.button("Add")
    # if add_clicked and dim_to_add and dim_to_add not in st.session_state.dims:
    #     st.session_state.dims.append(dim_to_add)

    # # Show selection / clear
    # st.write("Selected dimensions:", ", ".join(st.session_state.dims) if st.session_state.dims else "(none)")
    # if st.session_state.dims:
    #     dim_to_remove = st.selectbox(
    #         "Remove a selected dimension",
    #         options=st.session_state.dims,
    #         index=None,
    #         key="dim_remove"
    #     )
    #     remove_clicked = st.button("Remove selected")
    #     if remove_clicked and dim_to_remove:
    #         st.session_state.dims.remove(dim_to_remove)

    # # Clear all option for full reset
    # if st.button("Clear all selections"):
    #     st.session_state.dims = []

    selected_dims = st.multiselect(
        "Select grouping dimensions (optional, multi):",
        options=dim_candidates,
        default=[]  # Start empty
    )

    # Save the selection for use later (if you want to persist with session state)
    st.session_state.dims = selected_dims

    st.write("Selected dimensions:", ", ".join(selected_dims) if selected_dims else "(none)")



    # if st.button("Clear"):
    #     st.session_state.dims = []

    with st.form("build_table"):
        run = st.form_submit_button("Run")

    if not (st.session_state.dims):
        st.stop()

    dims = st.session_state.dims.copy()

    # -----------------------------
    # Build long form (melt)
    # -----------------------------
    id_vars = ["vjp_occasion_category"] + dims
    question_cols = filtered_df.columns[anchor+1:].tolist()
    # Safety: ensure id_vars not in question_cols
    for non_q in id_vars:
        if non_q in question_cols:
            question_cols.remove(non_q)

    df_long = filtered_df.melt(
        id_vars=id_vars,
        value_vars=question_cols,
        var_name="Question",
        value_name="Value"
    )

    # Yes/No -> 1/0; keep numeric as-is
    yn_map = df_long["Value"].astype("string").str.strip().str.lower().map({"yes": 1, "no": 0})
    numeric = pd.to_numeric(df_long["Value"], errors="coerce")
    df_long["Value"] = numeric.fillna(yn_map)

    # -----------------------------
    # Per-question total ("Total")
    # -----------------------------
    total = (df_long.groupby("Question", as_index=False)["Value"].sum()
                .rename(columns={"Value": "Total"})
                .set_index("Question"))

    # -----------------------------
    # Pivots (value-only labels; dedup across dims)
    # -----------------------------
    blocks = []
    for dim in dims:
        p = df_long.pivot_table(
            index="Question",
            columns=dim,
            values="Value",
            aggfunc="sum",
            fill_value=0
        )
        # Columns: use category values directly as labels
        p.columns = p.columns.astype(str)

        # De-duplicate labels across prior blocks if needed
        if blocks:
            existing = set().union(*[set(b.columns) for b in blocks])
            new_cols, counts = [], {}
            for c in p.columns:
                name = c
                if name in existing or name in new_cols:
                    counts[name] = counts.get(name, 1) + 1
                    name = f"{name} ({counts[name]})"
                new_cols.append(name)
            p.columns = new_cols

        blocks.append(p)

    out = pd.concat([total] + blocks, axis=1).reset_index()

    # -----------------------------
    # Add Total row (dynamic to dims)
    # -----------------------------
    def add_total_row_dynamic(out: pd.DataFrame, filtered_df: pd.DataFrame, dims: list) -> pd.DataFrame:
        counts = {d: filtered_df[d].value_counts() for d in dims}  # dropna default
        prefixed = any((":" in c) for c in out.columns if c not in ("Question", "Total"))

        total_row = {c: 0 for c in out.columns}
        total_row["Question"] = "Total"
        total_row["Total"] = int(len(filtered_df))

        for col in out.columns:
            if col in ("Question", "Total"):
                continue
            if prefixed and ":" in col:
                dim, cat = col.split(":", 1)
                if dim in counts:
                    total_row[col] = int(counts[dim].get(cat, 0))
            else:
                # match by category label only
                val = 0
                for d, s in counts.items():
                    if col in s.index:
                        val = int(s.get(col, 0))
                        break
                total_row[col] = val

        return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

    out = add_total_row_dynamic(out, filtered_df, dims)

    # Strip any prefix before '_' in Question
    out["Question"] = (
        out["Question"].astype("string").str.split("_", n=1).str[-1]
    )

    # -----------------------------
    # Shares and Index blocks
    # -----------------------------
    num_cols = [c for c in out.columns if c != "Question" and c.strip() != ""]
    mask_total = out["Question"].eq("Total")

    # Shares (% of column total), with Total row = column average
    totals = out.loc[mask_total, num_cols].squeeze()
    shares = out[num_cols].div(totals.replace(0, np.nan), axis="columns") * 100
    col_means = shares.loc[~mask_total, :].mean(axis=0)
    shares.loc[mask_total, :] = col_means.values

    share_block = shares.add_suffix("_share")
    blank1 = pd.DataFrame({"": [np.nan] * len(out)})
    out_with_shares = pd.concat([out, blank1, share_block], axis=1)

    # Index vs Total (row-wise), Total row blank
    share_cols = [c for c in out_with_shares.columns if c.endswith("_share")]
    shares_only = out_with_shares[share_cols]
    denom = shares_only["Total_share"].replace(0, np.nan)
    idx_vs_pw = shares_only.div(denom, axis="index") * 100
    idx_vs_pw.loc[mask_total, :] = np.nan
    idx_vs_pw = idx_vs_pw.add_suffix("_idx")

    blank2 = pd.DataFrame({"  ": [np.nan] * len(out_with_shares)})
    out_final = pd.concat([out_with_shares, blank2, idx_vs_pw], axis=1)
    # Drop redundant Total index column
    if "Total_share_idx" in out_final.columns:
        out_final.drop(columns=["Total_share_idx"], inplace=True)

    # -----------------------------
    # Display with % formatting
    # -----------------------------
    pct_cols = [c for c in out_final.columns if c.endswith("_share") or c.endswith("_idx")]
    col_cfg = {c: st.column_config.NumberColumn(label=c, format="%.0f %%") for c in pct_cols}
    # seg_value = ", ".join(custom_name)  # Combine all selected segments as a string

    out_final.insert(0, "Segment", [custom_name] * len(out_final))
    st.header("Summary output:")
    st.dataframe(out_final, use_container_width=True, column_config=col_cfg)

    # -------------------------------------------------------------
    # 📝 Shared Comment Box for Meastria
    # -------------------------------------------------------------

    # Create session storage if not already exists
    if "meastria_comment" not in st.session_state:
        st.session_state["meastria_comment"] = ""

    meastria_comment = st.text_area(
        "📝 Comment on Meastria Data:",
        value=st.session_state["meastria_comment"],
        key="meastria_comment_main"
    )

    # Update global storage
    st.session_state["meastria_comment"] = meastria_comment



    import streamlit as st
    import pandas as pd
    import openpyxl
    from io import BytesIO


    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font  # for bold styling
    import streamlit as st
    import pandas as pd
    import numpy as np

    # Preconditions:
    # - uploaded_file: the st.file_uploader handle
    # - st.session_state.selected_sheet: the sheet name you loaded from
    # - out_final: the DataFrame you want to export

    if uploaded_file is not None and "selected_sheet" in st.session_state and st.session_state.selected_sheet:
        base_sheet = st.session_state.selected_sheet
        summary_sheet = f"Summary_{base_sheet}"

        # Step 1: Reset and load the workbook
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file)

        # Step 2: Ensure all sheets are visible
        for ws in wb.worksheets:
            ws.sheet_state = "visible"

        # Step 3: Safely remove old summary if it exists
        if summary_sheet in wb.sheetnames:
            if wb.active.title == summary_sheet:
                wb.active = wb.worksheets[0]
            wb.remove(wb[summary_sheet])

        # Step 4: Create the new summary sheet directly via openpyxl
        ws_new = wb.create_sheet(title=summary_sheet)

        # Step 4a: Write headers (row 1) and make them bold
        for j, col_name in enumerate(out_final.columns, start=1):
            cell = ws_new.cell(row=1, column=j, value=str(col_name))
            cell.font = Font(bold=True)  # header bold [openpyxl Font] [web:425][web:441]

        # Step 4b: Write data (starting row 2), preserving raw numeric values
        for r_idx, row in enumerate(out_final.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_new.cell(row=r_idx, column=c_idx, value=value)  # keep exact numeric [web:425]

        # Step 4c: Apply non-scaling percent display for columns whose header contains "share"
        # Use Excel custom number format 0\% so 49.9556 displays as 50% but raw value stays 49.9556 [web:433][web:425]
        nrows = len(out_final)
        header_strs = [str(c) for c in out_final.columns]
        share_cols_idx = [i + 1 for i, name in enumerate(header_strs) if "share" in name.lower()]
        for col_idx in share_cols_idx:
            for r in range(2, nrows + 2):
                ws_new.cell(row=r, column=col_idx).number_format = r"0\%"

        # Step 4d: Bold all values in "Question" and "SOG" columns (entire columns, data rows only)
        # Resolve by header name; fallback to first two columns if names not found.
        name_to_idx = {str(name).strip(): i + 1 for i, name in enumerate(out_final.columns)}
        target_cols = []
        for key in ["Question", "Segment"]:
            if key in name_to_idx:
                target_cols.append(name_to_idx[key])
        if not target_cols:
            # Fallback: first two columns if headers not matched
            target_cols = [1, 2]
        for col_idx in target_cols:
            for r in range(2, nrows + 2):
                cell = ws_new.cell(row=r, column=col_idx)
                cell.font = Font(bold=True)  # column bold for values [web:425][web:441]

        # Optional: freeze top row
        ws_new.freeze_panes = "A2"  # header lock only, no autofilter [web:396]

        # Step 5: Save workbook to BytesIO buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Step 6: Provide download button
        st.download_button(
            label="Download Excel",
            data=output,
            file_name=f"updated_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )




with tab3:
    # st.title("📊 Manufacturing Company Analysis")

    if uploaded_file1 is not None:

        
        admix_filter = st.selectbox(
            "Select Segment",
            ["Admix Deluxe", "Admix Premium", "Admix Value", "S&PIB", "SP+IB"]
        )

        col1, col2 = st.columns(2)

        with col1:

            st.markdown(f"#### 🏭 Manufacturing Co. View")

            
            # Convert formatted columns back to numeric for plotting
            numeric_cols = ['A23', 'A24', 'A25']
            growth_cols = ['A24 Growth %', 'A25 Growth %']

            plot_df = final_mfs_review.copy()
            plot_df[numeric_cols] = plot_df[numeric_cols].replace({",": ""}, regex=True).astype(float)
            plot_df[growth_cols] = plot_df[growth_cols].replace("%", "", regex=True).astype(float)

            selected_year = st.selectbox("Select PRI Year for Sales Pie Chart", options=numeric_cols)

            import plotly.express as px

            pie_colors = {
                "PRI": "#9c7ad4",       # Blue
                "Diageo": "#6d41a7",    # Orange
                "Others": "#4d1780"     # Green
            }


            # pie_fig = px.pie(
            #     plot_df,
            #     names='Mfg Com',
            #     values=selected_year,
            #     title=f"Manufacturing Company NS Share ({selected_year})",
            #     hole=0.4,
            #     color='Mfg Com',
            #     color_discrete_map=pie_colors
            # )

            # pie_fig.update_layout(
            #     width=380,
            #     height=350
            # )

            # st.plotly_chart(pie_fig, use_container_width=True)

            lollipop_df = plot_df.sort_values(by=selected_year)

            lollipop_fig = px.scatter(
                lollipop_df,
                x=selected_year,
                y="Mfg Com",
                size=[10]*len(lollipop_df),
                color="Mfg Com",
                color_discrete_map=pie_colors
            )

            for _, row in lollipop_df.iterrows():
                lollipop_fig.add_shape(
                    type="line",
                    x0=0, x1=row[selected_year],
                    y0=row["Mfg Com"], y1=row["Mfg Com"]
                )

            lollipop_fig.update_layout(
                title=f"NS Share Comparison ({selected_year})",
                width=380,
                height=350,
                showlegend=False
            )

            st.plotly_chart(lollipop_fig, use_container_width=True)





            # with st.expander("Add comments:"):

            #     # ---- Comment box for PIE chart ----
            #     pie_comment = st.text_area(
            #         "📝 Comment on NS Share:",
            #         placeholder="Example: PRI leads NS share in A25, competitors losing share...",
            #         key="pie_comment"
            #     )

            #     if pie_comment:
            #         st.caption("✔ Comment saved!")

        # with col2:

            # growth_melt = plot_df.melt(
            #     id_vars='Mfg Com',
            #     value_vars=growth_cols,
            #     var_name='Year',
            #     value_name='Growth %'
            # )

            # # Clean label formatting
            # growth_melt['Year'] = growth_melt['Year'].str.replace(" Growth %", "")

            # bar_fig = px.bar(
            #     growth_melt,
            #     x='Mfg Com',
            #     y='Growth %',
            #     color='Year',
            #     barmode='group',
            #     text='Growth %',
            #     title="YoY Growth % by Manufacturing Company"
            # )
            # bar_fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
            # bar_fig.update_layout(yaxis_title="Growth (%)")
            # bar_fig.update_layout(
            #     width=500,
            #     height=380
            # )


            # st.plotly_chart(bar_fig, use_container_width=True)

            import plotly.graph_objects as go
            import numpy as np

            # st.markdown("### 📊 Brand Family NS (Mn INR) – A23 vs A24 vs A25")

            # -----------------------------
            # Prepare Data
            # -----------------------------

            pivot_brand_family= pd.pivot_table( df_filtered, index=['Brand Family'], columns=['PRI Year'], values='NS M INR', aggfunc='sum', fill_value=0 ).sort_index(axis=1, level=[0, 1])

            pivot = pivot_brand_family.copy()
            pivot = pivot[['A23', 'A24', 'A25']]  # ensure correct order

            # Separate Blenders Pride
            bp_row = pivot.loc[pivot.index == "Blenders Pride"]

            # Other brands sorted by A25 descending
            other_rows = pivot.loc[pivot.index != "Blenders Pride"] \
                            .sort_values(by="A25", ascending=False)

            # Final ordered dataframe
            pivot = pd.concat([bp_row, other_rows])

            brand_families = pivot.index.tolist()
            a23 = pivot['A23'].values
            a24 = pivot['A24'].values
            a25 = pivot['A25'].values

            # -----------------------------
            # YoY Growth Calculations
            # -----------------------------

            yoy_24 = np.where(a23 > 0, (a24 - a23) / a23 * 100, 0)
            yoy_25 = np.where(a24 > 0, (a25 - a24) / a24 * 100, 0)

            # -----------------------------
            # CAGR (2 Year)
            # -----------------------------

            cagr_2yr = np.where(a23 > 0, ((a25 / a23) ** (1 / 2) - 1) * 100, 0)

            # -----------------------------
            # Plotly Bar Chart
            # -----------------------------

            colors = {
                "A23": "#62b6f1",   # Blue
                "A24": "#1e94d8",   # Orange
                "A25": "#1b4ebd"    # Green
            }

            fig = go.Figure()

            fig.add_bar(name="A23", x=brand_families, y=a23, marker_color=colors["A23"])
            fig.add_bar(name="A24", x=brand_families, y=a24, marker_color=colors["A24"])
            fig.add_bar(name="A25", x=brand_families, y=a25, marker_color=colors["A25"])
            # -----------------------------
            # Annotations (YoY & CAGR)
            # -----------------------------

            for i, bf in enumerate(brand_families):
                fig.add_annotation(
                    x=bf, y=a24[i],
                    text=f"+{yoy_24[i]:.1f}%",
                    showarrow=False,
                    yshift=15
                )

                fig.add_annotation(
                    x=bf, y=a25[i],
                    text=f"+{yoy_25[i]:.1f}%",
                    showarrow=False,
                    yshift=15
                )

                fig.add_annotation(
                    x=bf, y=max(a25[i], a24[i], a23[i]),
                    text=f"CAGR (2yr): {cagr_2yr[i]:.1f}%",
                    showarrow=False,
                    yshift=35
                )

            # -----------------------------
            # Layout
            # -----------------------------

            fig.update_layout(
                barmode="group",
                xaxis_title="Brand Family",
                yaxis_title="NS (Mn INR)",
                height=450,
                legend_title="Year",
                title= "Brand Family NS (Mn INR) – A23 vs A24 vs A25"
            )

            st.plotly_chart(fig, use_container_width=True)


            # with st.expander("Add comments:"):

            #     # ---- Comment box for BAR chart ----
            #     bar_comment = st.text_area(
            #         "📝 Comment on Growth Trend:",
            #         placeholder="Example: PRI growth slowing YoY while Diageo improving...",
            #         key="bar_comment"
            #     )

            #     if bar_comment:
            #         st.caption("✔ Comment saved!")

        

        with col2:
            import streamlit as st
            import pandas as pd
            import json
            import requests
            import plotly.express as px

            st.markdown(f"#### 📍 India State Category Map (P1 / P2 / Others)")


            # --------------------------
            # 1) Load India States GeoJSON
            # --------------------------
            geojson_url = "https://raw.githubusercontent.com/geohacker/india/master/state/india_telengana.geojson"
            india_geo = requests.get(geojson_url).json()

            # geojson_url = "https://raw.githubusercontent.com/Anujarya300/bihar_pollution/master/data/india_states.geojson"
            # resp = requests.get(geojson_url)
            # india_geo = resp.json()


            # --------------------------
            # 2) Define Dry States
            # --------------------------
            dry_states = ["Bihar", "Gujarat", "Mizoram", "Nagaland"]

            # --------------------------
            # 3) Prepare Pivot Table Filter
            # --------------------------

            pivot_share_no_total = pivot_share.drop(index="Seg Total", errors="ignore")

            if "A25" not in pivot_share_no_total.columns:
                st.error("⚠ A25 column not found in pivot_share.")
            else:
                share_df = pivot_share_no_total["A25"].reset_index()
                share_df.columns = ["state", "share_A25"]

            # --------------------------
            # 4) Assign Dry States Category
            # --------------------------
            # share_df = share_df[~share_df["state"].isin(dry_states)]

            share_df["category"] = share_df["state"].apply(lambda s: "Dry State" if s in dry_states else None)

            # --------------------------
            # 5) Apply P1 & P2 Logic (excluding Dry States)
            # --------------------------

            ranked_df = share_df[share_df["category"].isna()].copy()
            ranked_df = ranked_df.sort_values(by="share_A25", ascending=False).reset_index(drop=True)

            ranked_df["rank"] = ranked_df.index + 1
            ranked_df["category"] = ranked_df["rank"].apply(
                lambda x: "P1" if x <= 5 else ("P2" if x <= 10 else "Others")
            )

            # --------------------------
            # 6) Merge categories back
            # --------------------------

            share_df = share_df.drop(columns=["category"])
            share_df = pd.merge(share_df, ranked_df[["state", "category"]], on="state", how="left")
            # share_df["category"] = share_df["state"].apply(lambda s: "Dry State" if s in dry_states else None)

            # Assign Dry state category back
            share_df.loc[share_df["state"].isin(dry_states), "category"] = "Dry State"

            # Ensure Bihar exists
            if "Bihar" not in share_df["state"].values:
                share_df = pd.concat([
                    share_df,
                    pd.DataFrame({"state": ["Bihar"], "share_A25": 0, "category": "Dry State"})
                ], ignore_index=True)

            # st.write(share_df["state"].unique())

            # --------------------------
            # 7) Fix naming mismatch for geojson
            # --------------------------

            name_mapping = {
                # "Orissa": "Odisha",
                "Tamilnadu": "Tamil Nadu",
                "Uttaranchal": "Uttarakhand",
                "Chattisgarh": "Chhattisgarh",
                "Jammu & Kashmir": "Jammu and Kashmir",
                "Pondicherry": "Puducherry",
                "Yanam": "Puducherry",
                "Mahe": "Puducherry",
                "Daman": "Dadra and Nagar Haveli and Daman and Diu",
                "Diu": "Dadra and Nagar Haveli and Daman and Diu",
                "Silvassa, Dadra And Nagar Haveli": "Dadra and Nagar Haveli and Daman and Diu",
                "Andaman Nicobar Island": "Andaman and Nicobar Islands",
                "Delhi": "Delhi"
            }

            share_df["state"] = share_df["state"].replace(name_mapping)

            # --------------------------
            # 8) Ensure ALL India map states exist (excluding dry states)
            # --------------------------

            dry_states = ["Bihar", "Gujarat", "Mizoram", "Nagaland"]

            geo_states = [feature["properties"]["NAME_1"] for feature in india_geo["features"]]

            # Remove dry states from the list of geo states to check
            states_to_check = [s for s in geo_states if s not in dry_states]

            # Find missing (only non-dry)
            missing_states = list(set(states_to_check) - set(share_df["state"]))

            if missing_states:
                # st.warning(f"🧩 Adding missing NON-DRY states: {missing_states}")
                share_df = pd.concat([
                    share_df,
                    pd.DataFrame({
                        "state": missing_states,
                        "share_A25": 0,
                        "category": "Others"
                    })
                ], ignore_index=True)
            else:
                st.success("✅ No non-dry states missing.")


            # share_df["category"] = share_df["state"].apply(lambda s: "Dry State" if s in dry_states else None)

            # --------------------------
            # 9) Final Output
            # --------------------------

            df_states = share_df.copy()

            # st.subheader("🔥 Final State Categorization (Including Dry States & Map Sync)")
            # st.dataframe(df_states)


        

            # --------------------------
            # 3) Plot Choropleth Map
            # --------------------------

            fig = px.choropleth(
                df_states,
                geojson=india_geo,
                locations="state",
                featureidkey="properties.NAME_1",
                color="category",
                title="India Map — P1 | P2 | Others",
                color_discrete_map={
                    "P1": "#2ecc71",     # Green
                    "P2": "#f39c12",     # Orange
                    "Others": "#bdc3c7"  # Grey
                },
            )

            # ✅ Make the map fit the states nicely
            fig.update_geos(
                fitbounds="locations",
                visible=False
            )

            # ✅ Increase figure height + add top margin space
            fig.update_layout(
                height=750,   # 👈 INCREASE graph height (try 700–900)
                margin={
                    "r": 0,
                    "t": 90,    # 👈 MORE SPACE ABOVE TITLE
                    "l": 0,
                    "b": 10
                },
                title={
                    "x": 0.5,      # center title
                    "xanchor": "center",
                    "font": {
                        "size": 20  # slightly bigger title
                    }
                }
            )

            st.plotly_chart(fig, use_container_width=True)

            st.markdown(
                """
                <p style="font-style: italic; font-size: 16px; color: #444; margin-top: 10px;">
                    P1: High Contribution States with High to Moderate Growth<br>
                    P2: Mod Contribution States with High to Moderate Growth
                </p>
                """,
                unsafe_allow_html=True
            )

        # ---- Comment box for BAR chart ----
        bar_comment = st.text_area(
            "📝 Comment:",
            placeholder="Example: P1 states contribute significantly to A25 share...",
            key="India_map_comment"
        )

        if bar_comment:
            st.caption("✔ Comment saved!")


        st.markdown(
            """ 
            <div style="height: 2px; background-color: black; margin: 15px 0;"></div>
            """, 
            unsafe_allow_html=True
            )

        col3, col4 = st.columns(2)

        with col3:

            import plotly.graph_objects as go
            import streamlit as st

            # -----------------------------
            # DATA
            # -----------------------------

            quarters = ["Q1", "Q2", "Q3", "Q4"]

            a24 = [12.5, 18.7, -6.3, -18.3]
            a25 = [29.5, 5.6, -11.9, -2.5]

            # -----------------------------
            # PLOT
            # -----------------------------

            fig1 = go.Figure()

            fig1.add_trace(go.Scatter(
                x=quarters, y=a24,
                mode="lines+markers+text",
                name="A24",
                text=[f"{v}%" for v in a24],
                textposition="top center"
            ))

            fig1.add_trace(go.Scatter(
                x=quarters, y=a25,
                mode="lines+markers+text",
                name="A25",
                text=[f"{v}%" for v in a25],
                textposition="top center"
            ))

            # -----------------------------
            # Q3 Highlight (Dashed Circle)
            # -----------------------------

            fig1.add_shape(
                type="circle",
                xref="x", yref="y",
                x0=1.6, x1=2.4,
                y0=-20, y1=5,
                line=dict(dash="dash")
            )

            # -----------------------------
            # Layout
            # -----------------------------

            fig1.update_layout(
                title="Premium Whisky Q-o-Q NS Gr % (BTM)",
                xaxis_title="Quarter",
                yaxis_title="Growth %",
                height=450,
                legend_title="Year"
            )

            st.plotly_chart(fig1, use_container_width=True)

        with col4:

            # -----------------------------
            # DATA
            # -----------------------------

            a25_ns = [29.5, 5.6, -11.9, -2.5]
            a25_p3m = [3.3, 3.7, -0.4, 8.1]

            # -----------------------------
            # PLOT
            # -----------------------------

            fig2 = go.Figure()

            fig2.add_trace(go.Scatter(
                x=quarters, y=a25_ns,
                mode="lines+markers+text",
                name="A25 NS",
                text=[f"{v}%" for v in a25_ns],
                textposition="top center"
            ))

            fig2.add_trace(go.Scatter(
                x=quarters, y=a25_p3m,
                mode="lines+markers+text",
                name="A25 P3M",
                text=[f"{v}%" for v in a25_p3m],
                textposition="top center"
            ))

            # -----------------------------
            # Q3 Highlight (Dashed Circle)
            # -----------------------------

            fig2.add_shape(
                type="circle",
                xref="x", yref="y",
                x0=1.6, x1=2.4,
                y0=-20, y1=8,
                line=dict(dash="dash")
            )

            # -----------------------------
            # Layout
            # -----------------------------

            fig2.update_layout(
                title="Premium Whisky Q-o-Q NS vs P3M Gr %",
                xaxis_title="Quarter",
                yaxis_title="Growth %",
                height=450,
                legend_title="Metric"
            )

            st.plotly_chart(fig2, use_container_width=True)

        st.markdown(
            """ 
            <div style="height: 2px; background-color: black; margin: 15px 0;"></div>
            """, 
            unsafe_allow_html=True
            )
        
        with st.expander("Summary Tabs"):

            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                "Performance",
                "State - Salience",
                "Profile",
                "Segment Truths",
                "Macro Indicators"
            ])

            with tab1:
                st.markdown(f"##### 🏭 Manufacturing Co. View")
                st.dataframe(final_mfs_review)

                # Show the same comment
                st.text_area(
                    "📝 Comment on Manufacturing Company View:",
                    value=st.session_state.get("mfg_table_comment", ""),
                    key="mfg_review_comment_tab"
                )

                # option = st.selectbox(
                #     "Display Options",
                #     ["Hide Brand Summary", "Show Brand Summary"]
                # )

                # # if option == "Show Brand Summary":
                # with st.expander("Brand-Level Summary (Click to Expand)"):
                #     st.subheader(f"🏭 Brand × Manufacturing Company — NS (₹ Mn) & YoY Growth — {selected_seg}")
                #     st.dataframe(styled_display)

                # --- Display Monthly Pivot ---
                st.markdown(f"##### 📅 Monthly Data (NS)")
                st.dataframe(combined_pivot.round(0).astype(int))

                st.text_area(
                    "📝 Comment on Monthly Data (NS):",
                    value=st.session_state.get("monthly_pivot_comment", ""),
                    key="monthly_pivot_comment_copy"
                )



            with tab2:

                # --- Step 6: Display ---
                st.markdown(f"##### 📊 Segment, Brand & Growth Comparison")
                st.dataframe(final_display_df.style.format("{:.2f}%"))

                st.text_area(
                    "📝 Comment on Segment, Brand & Growth Comparison:",
                    value=st.session_state.get("seg_brand_growth_comment", ""),
                    key="seg_brand_growth_comment_box_duplicate"
                )

            with tab3:

                import streamlit as st
                import pandas as pd

                # -----------------------------
                # Create the DataFrame
                # -----------------------------

                data = {
                    "P3M Seg Profile": [
                        "LDA-24",
                        "25-30",
                        "LDA-35",
                        "36-45",
                        "46+",
                        "NCCS A",
                        "Single",
                        "Married w/ Kids",
                        "Married w/o Kids",
                        "Single Parent",
                        "Business Owners",
                        "Salaried",
                        "High",
                        "Medium",
                        "Low",
                        "NE: Directs Entrant",
                        "NE: From Beer",
                        "NE: From Whites",
                        "NE: NETT",
                        "Non-Entrant"
                    ],
                    "TBA": [
                        "8%",
                        "43%",
                        "56%",
                        "32%",
                        "17%",
                        "70",
                        "28%",
                        "41%",
                        "9%",
                        "21%",
                        "40%",
                        "56%",
                        "63%",
                        "28%",
                        "9%",
                        "9%",
                        "8%",
                        "10%",
                        "27%",
                        "73%"
                    ],
                    "Premium Whisky": [
                        "10%",
                        "45%",
                        "60%",
                        "31%",
                        "14%",
                        "70",
                        "32%",
                        "34%",
                        "7%",
                        "26%",
                        "41%",
                        "56%",
                        "68%",
                        "25%",
                        "6%",
                        "10%",
                        "9%",
                        "13%",
                        "32%",
                        "68%"
                    ]
                }

                df = pd.DataFrame(data)

                # -----------------------------
                # Streamlit UI
                # -----------------------------

                # st.set_page_config(page_title="Consumer Profile Dashboard", layout="wide")

                st.markdown("##### 📊 P3M Segment Profile — TBA vs Premium Whisky")

                st.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True
                )


            with tab4:

                # st.subheader("Linked with Meastria")
                st.markdown("##### Meastria Summary output:")
                st.dataframe(out_final, use_container_width=True, column_config=col_cfg)

                st.text_area(
                    "📝 Comment on Meastria Data:",
                    value=st.session_state.get("meastria_comment", ""),
                    key="meastria_comment_copy"
                )

            with tab5:

                # -----------------------------
                # P1 & P2 Macro Ranks (Updated with Your Data)
                # -----------------------------

                macro_data = {
                    "Priority": ["P1", "P2", "P2", "P1", "P1", "P1", "P2", "P2", "P1", "P2"],
                    "State": [
                        "Maharashtra",
                        "Karnataka",
                        "Haryana",
                        "Uttar Pradesh",
                        "Telangana",
                        "West Bengal",
                        "Odisha",
                        "Rajasthan",
                        "Andhra Pradesh",
                        "Assam"
                    ],
                    "GDP": [1, 3, 11, 4, 7, 5, 13, 6, 8, 15],
                    "MFG": [1, 4, 8, 3, 11, 10, 5, 7, 6, 18],
                    "LFP": [6, 15, 4, 19, 2, 7, 12, 3, 9, 10],
                    "LFP-Growth": [20, 21, 24, 5, 23, 25, 18, 4, 9, 6],
                    "GST": [1, 2, 4, 5, 7, 8, 9, 10, 12, 18]
                }

                df_macro = pd.DataFrame(macro_data)

                # -----------------------------
                # Display Macro Table
                # -----------------------------

                st.markdown("##### 📌 P1 & P2 Overview — Macro Ranks (State-wise)")

                st.dataframe(
                    df_macro,
                    use_container_width=True,
                    hide_index=True
                )
        st.markdown(
            """ 
            <div style="height: 3px; background-color: black; margin: 15px 0;"></div>
            """, 
            unsafe_allow_html=True
            )







    else:
        st.info("Please upload an Excel file to continue.")


    
    


        

