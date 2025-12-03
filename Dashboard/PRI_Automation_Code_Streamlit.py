


########################## Correct code with flexibility(17102025) ######################### some raw files had less column so to make it user friendly and also the responses were in Yes and No instead of 1 and 0


import streamlit as st
import pandas as pd
import numpy as np

# -----------------------------
# Page setup
# -----------------------------
st.set_page_config(page_title="PRI Automation", layout="wide")
st.title("📊 PRI Maestria Data Tool")
st.header("Upload Input Files")

# -----------------------------
# Session state init
# -----------------------------
for key, default in {
    "submitted": False,
    "selected_sheet": None,
    "prev_file": None,
    "df": None,
    "dims": []
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# -----------------------------
# File upload + sheet selection (gated)
# -----------------------------
uploaded_file = st.file_uploader("Upload the Excel Data File", type=["xlsx"], key="excel_file")

# Reset state if file changes
cur_file = getattr(uploaded_file, "name", None)
if cur_file != st.session_state.prev_file:
    st.session_state.submitted = False
    st.session_state.selected_sheet = None
    st.session_state.prev_file = cur_file
    st.session_state.df = None  # reset cached df too

if uploaded_file is None:
    st.info("Upload a file to begin.")
    st.stop()

xls = pd.ExcelFile(uploaded_file)
with st.form("sheet_picker_form", clear_on_submit=False):
    sheet = st.selectbox(
        "Select a sheet",
        options=xls.sheet_names,
        index=None,
        placeholder="Choose a sheet…",
        key="sheet_name"
    )
    clicked = st.form_submit_button("Load sheet", type="primary")

if clicked and sheet is not None:
    st.session_state.submitted = True
    st.session_state.selected_sheet = sheet

if not (st.session_state.submitted and st.session_state.selected_sheet):
    st.info("Select a sheet and click Load sheet to proceed.")
    st.stop()

# Load df once per file/sheet
if st.session_state.df is None:
    df = pd.read_excel(uploaded_file, sheet_name=st.session_state.selected_sheet)
    st.session_state.df = df.copy()
else:
    df = st.session_state.df

st.success(f"Loaded sheet: {st.session_state.selected_sheet}")
# Optional: show raw df preview (comment out if not needed)
st.dataframe(df, use_container_width=True)

# -----------------------------
# Segment selection
# -----------------------------
# Build segment list safely
if "vjp_occasion_category" not in df.columns:
    st.error("Column 'vjp_occasion_category' is missing in this sheet.")
    st.stop()

segment_options = (
    df["vjp_occasion_category"].dropna().unique().tolist()
)
segment_options = sorted(segment_options)

st.header("Select segments:")
selected_seg = st.multiselect(
    "Select Revised Segment(s):",
    options=segment_options,
    default=[]
    # default=segment_options[:1],  # optional, default selection
)

if not selected_seg:
    st.warning("Please select at least one segment to continue.")
    st.stop()

# filtered_df = df[df['vjp_occasion_category'].isin(selected_seg)]

# 2. Specify custom name for selected segments
default_name = selected_seg[0] if len(selected_seg) == 1 else " + ".join(selected_seg)
custom_name = st.text_input(
    "Enter a new label for the segment(s):",
    value=default_name
)

# 3. Rename selected segments in the full DataFrame
df_renamed = df.copy()
df_renamed.loc[
    df_renamed["vjp_occasion_category"].isin(selected_seg), "vjp_occasion_category"
] = custom_name

filtered_df = df_renamed[df_renamed["vjp_occasion_category"] == custom_name]

st.header("Filtered Data:")
st.dataframe(filtered_df, use_container_width=True)

# -----------------------------
# Dimension picking (single-add) + run
# -----------------------------
# Guard for empty filtered_df
if filtered_df.empty:
    st.warning("No rows for the selected segment.")
    st.stop()

# Anchor column must exist
if "QAge" not in filtered_df.columns:
    st.error("Anchor column 'QAge' is missing.")
    st.stop()

anchor = filtered_df.columns.get_loc("QAge")
dim_candidates = [c for c in filtered_df.columns[:anchor+1] if c != "vjp_occasion_category"]

# Single add UI
st.header("Add grouping dimensions:")
# dim_to_add = st.selectbox(
#     "Add a grouping dimension",
#     options=[d for d in dim_candidates if d not in st.session_state.dims],
#     index=None
# )
# add_clicked = st.button("Add")
# if add_clicked and dim_to_add and dim_to_add not in st.session_state.dims:
#     st.session_state.dims.append(dim_to_add)

# # Show selection / clear
# st.write("Selected dimensions:", ", ".join(st.session_state.dims) if st.session_state.dims else "(none)")
# if st.session_state.dims:
#     dim_to_remove = st.selectbox(
#         "Remove a selected dimension",
#         options=st.session_state.dims,
#         index=None,
#         key="dim_remove"
#     )
#     remove_clicked = st.button("Remove selected")
#     if remove_clicked and dim_to_remove:
#         st.session_state.dims.remove(dim_to_remove)

# # Clear all option for full reset
# if st.button("Clear all selections"):
#     st.session_state.dims = []

selected_dims = st.multiselect(
    "Select grouping dimensions (optional, multi):",
    options=dim_candidates,
    default=[]  # Start empty
)

# Save the selection for use later (if you want to persist with session state)
st.session_state.dims = selected_dims

st.write("Selected dimensions:", ", ".join(selected_dims) if selected_dims else "(none)")



# if st.button("Clear"):
#     st.session_state.dims = []

with st.form("build_table"):
    run = st.form_submit_button("Run")

if not (run and st.session_state.dims):
    st.stop()

dims = st.session_state.dims.copy()

# -----------------------------
# Build long form (melt)
# -----------------------------
id_vars = ["vjp_occasion_category"] + dims
question_cols = filtered_df.columns[anchor+1:].tolist()
# Safety: ensure id_vars not in question_cols
for non_q in id_vars:
    if non_q in question_cols:
        question_cols.remove(non_q)

df_long = filtered_df.melt(
    id_vars=id_vars,
    value_vars=question_cols,
    var_name="Question",
    value_name="Value"
)

# Yes/No -> 1/0; keep numeric as-is
yn_map = df_long["Value"].astype("string").str.strip().str.lower().map({"yes": 1, "no": 0})
numeric = pd.to_numeric(df_long["Value"], errors="coerce")
df_long["Value"] = numeric.fillna(yn_map)

# -----------------------------
# Per-question total ("Total")
# -----------------------------
total = (df_long.groupby("Question", as_index=False)["Value"].sum()
              .rename(columns={"Value": "Total"})
              .set_index("Question"))

# -----------------------------
# Pivots (value-only labels; dedup across dims)
# -----------------------------
blocks = []
for dim in dims:
    p = df_long.pivot_table(
        index="Question",
        columns=dim,
        values="Value",
        aggfunc="sum",
        fill_value=0
    )
    # Columns: use category values directly as labels
    p.columns = p.columns.astype(str)

    # De-duplicate labels across prior blocks if needed
    if blocks:
        existing = set().union(*[set(b.columns) for b in blocks])
        new_cols, counts = [], {}
        for c in p.columns:
            name = c
            if name in existing or name in new_cols:
                counts[name] = counts.get(name, 1) + 1
                name = f"{name} ({counts[name]})"
            new_cols.append(name)
        p.columns = new_cols

    blocks.append(p)

out = pd.concat([total] + blocks, axis=1).reset_index()

# -----------------------------
# Add Total row (dynamic to dims)
# -----------------------------
def add_total_row_dynamic(out: pd.DataFrame, filtered_df: pd.DataFrame, dims: list) -> pd.DataFrame:
    counts = {d: filtered_df[d].value_counts() for d in dims}  # dropna default
    prefixed = any((":" in c) for c in out.columns if c not in ("Question", "Total"))

    total_row = {c: 0 for c in out.columns}
    total_row["Question"] = "Total"
    total_row["Total"] = int(len(filtered_df))

    for col in out.columns:
        if col in ("Question", "Total"):
            continue
        if prefixed and ":" in col:
            dim, cat = col.split(":", 1)
            if dim in counts:
                total_row[col] = int(counts[dim].get(cat, 0))
        else:
            # match by category label only
            val = 0
            for d, s in counts.items():
                if col in s.index:
                    val = int(s.get(col, 0))
                    break
            total_row[col] = val

    return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)

out = add_total_row_dynamic(out, filtered_df, dims)

# Strip any prefix before '_' in Question
out["Question"] = (
    out["Question"].astype("string").str.split("_", n=1).str[-1]
)

# -----------------------------
# Shares and Index blocks
# -----------------------------
num_cols = [c for c in out.columns if c != "Question" and c.strip() != ""]
mask_total = out["Question"].eq("Total")

# Shares (% of column total), with Total row = column average
totals = out.loc[mask_total, num_cols].squeeze()
shares = out[num_cols].div(totals.replace(0, np.nan), axis="columns") * 100
col_means = shares.loc[~mask_total, :].mean(axis=0)
shares.loc[mask_total, :] = col_means.values

share_block = shares.add_suffix("_share")
blank1 = pd.DataFrame({"": [np.nan] * len(out)})
out_with_shares = pd.concat([out, blank1, share_block], axis=1)

# Index vs Total (row-wise), Total row blank
share_cols = [c for c in out_with_shares.columns if c.endswith("_share")]
shares_only = out_with_shares[share_cols]
denom = shares_only["Total_share"].replace(0, np.nan)
idx_vs_pw = shares_only.div(denom, axis="index") * 100
idx_vs_pw.loc[mask_total, :] = np.nan
idx_vs_pw = idx_vs_pw.add_suffix("_idx")

blank2 = pd.DataFrame({"  ": [np.nan] * len(out_with_shares)})
out_final = pd.concat([out_with_shares, blank2, idx_vs_pw], axis=1)
# Drop redundant Total index column
if "Total_share_idx" in out_final.columns:
    out_final.drop(columns=["Total_share_idx"], inplace=True)

# -----------------------------
# Display with % formatting
# -----------------------------
pct_cols = [c for c in out_final.columns if c.endswith("_share") or c.endswith("_idx")]
col_cfg = {c: st.column_config.NumberColumn(label=c, format="%.0f %%") for c in pct_cols}
# seg_value = ", ".join(custom_name)  # Combine all selected segments as a string

out_final.insert(0, "Segment", [custom_name] * len(out_final))
st.header("Summary output:")
st.dataframe(out_final, use_container_width=True, column_config=col_cfg)


import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

# if uploaded_file is not None and "selected_sheet" in st.session_state and st.session_state.selected_sheet:
#     base_sheet = st.session_state.selected_sheet
#     summary_sheet = f"Summary_{base_sheet}"

#     # Step 1: Reset and load the workbook
#     uploaded_file.seek(0)
#     wb = openpyxl.load_workbook(uploaded_file)

#     # Step 2: Ensure all sheets are visible
#     for ws in wb.worksheets:
#         ws.sheet_state = "visible"

#     # Step 3: Safely remove old summary if it exists
#     if summary_sheet in wb.sheetnames:
#         if wb.active.title == summary_sheet:
#             wb.active = 0
#         wb.remove(wb[summary_sheet])

#     # Step 4: Create the new summary sheet directly via openpyxl
#     ws_new = wb.create_sheet(title=summary_sheet)

#     # Write DataFrame to that sheet manually
#     for r_idx, row in enumerate(out_final.itertuples(index=False), start=2):
#         for c_idx, value in enumerate(row, start=1):
#             ws_new.cell(row=r_idx, column=c_idx, value=value)

#     # Write headers
#     for i, col_name in enumerate(out_final.columns, start=1):
#         ws_new.cell(row=1, column=i, value=col_name)

#     # Step 5: Save workbook to BytesIO buffer
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     # Step 6: Provide download button
#     st.download_button(
#         label="Download updated workbook",
#         data=output,
#         file_name=f"updated_{uploaded_file.name}",
#         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#     )

from io import BytesIO
import openpyxl
from openpyxl.styles import Font  # for bold styling
import streamlit as st
import pandas as pd
import numpy as np

# Preconditions:
# - uploaded_file: the st.file_uploader handle
# - st.session_state.selected_sheet: the sheet name you loaded from
# - out_final: the DataFrame you want to export

if uploaded_file is not None and "selected_sheet" in st.session_state and st.session_state.selected_sheet:
    base_sheet = st.session_state.selected_sheet
    summary_sheet = f"Summary_{base_sheet}"

    # Step 1: Reset and load the workbook
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(uploaded_file)

    # Step 2: Ensure all sheets are visible
    for ws in wb.worksheets:
        ws.sheet_state = "visible"

    # Step 3: Safely remove old summary if it exists
    if summary_sheet in wb.sheetnames:
        if wb.active.title == summary_sheet:
            wb.active = wb.worksheets[0]
        wb.remove(wb[summary_sheet])

    # Step 4: Create the new summary sheet directly via openpyxl
    ws_new = wb.create_sheet(title=summary_sheet)

    # Step 4a: Write headers (row 1) and make them bold
    for j, col_name in enumerate(out_final.columns, start=1):
        cell = ws_new.cell(row=1, column=j, value=str(col_name))
        cell.font = Font(bold=True)  # header bold [openpyxl Font] [web:425][web:441]

    # Step 4b: Write data (starting row 2), preserving raw numeric values
    for r_idx, row in enumerate(out_final.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_new.cell(row=r_idx, column=c_idx, value=value)  # keep exact numeric [web:425]

    # Step 4c: Apply non-scaling percent display for columns whose header contains "share"
    # Use Excel custom number format 0\% so 49.9556 displays as 50% but raw value stays 49.9556 [web:433][web:425]
    nrows = len(out_final)
    header_strs = [str(c) for c in out_final.columns]
    share_cols_idx = [i + 1 for i, name in enumerate(header_strs) if "share" in name.lower()]
    for col_idx in share_cols_idx:
        for r in range(2, nrows + 2):
            ws_new.cell(row=r, column=col_idx).number_format = r"0\%"

    # Step 4d: Bold all values in "Question" and "SOG" columns (entire columns, data rows only)
    # Resolve by header name; fallback to first two columns if names not found.
    name_to_idx = {str(name).strip(): i + 1 for i, name in enumerate(out_final.columns)}
    target_cols = []
    for key in ["Question", "Segment"]:
        if key in name_to_idx:
            target_cols.append(name_to_idx[key])
    if not target_cols:
        # Fallback: first two columns if headers not matched
        target_cols = [1, 2]
    for col_idx in target_cols:
        for r in range(2, nrows + 2):
            cell = ws_new.cell(row=r, column=col_idx)
            cell.font = Font(bold=True)  # column bold for values [web:425][web:441]

    # Optional: freeze top row
    ws_new.freeze_panes = "A2"  # header lock only, no autofilter [web:396]

    # Step 5: Save workbook to BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Step 6: Provide download button
    st.download_button(
        label="Download Excel",
        data=output,
        file_name=f"updated_{uploaded_file.name}",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

